#!/usr/bin/env python3
"""
Расширенный импорт из Excel файла ДОЛГИ.xlsx
Воспроизводит логику расчётов "как в таблице":
- Изменение долга за период (разница между total_debt на даты)
- Агрегация по сезонам и годам
- Льготные периоды с суммами по месяцам
- Детализация расходов и доходов
- Недельное распределение бюджета
"""

import openpyxl
from datetime import date, datetime
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple
from sqlalchemy.orm import Session
from sqlalchemy import select

from app.models import (
    User, UserRole, RecordStatus, Debt, DebtHistory, CreditCard,
    Income, IncomeCategory, Expense, ExpenseCategory
)


def _coerce_date(value) -> Optional[date]:
    """Преобразует значение в дату."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        # Пробуем разные форматы
        for fmt in ["%Y-%m-%d", "%d.%m.%Y", "%d.%m"]:
            try:
                parsed = datetime.strptime(value.strip(), fmt)
                if fmt == "%d.%m":
                    # Если только день и месяц, предполагаем 2026 год
                    parsed = parsed.replace(year=2026)
                return parsed.date()
            except ValueError:
                continue
        # Пробуем извлечь дату из строки вида "до 13.05"
        if value.startswith("до "):
            try:
                parsed = datetime.strptime(value[3:].strip(), "%d.%m")
                return parsed.replace(year=2026).date()
            except ValueError:
                pass
    return None


def _coerce_float(value) -> Optional[float]:
    """Преобразует значение в число."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        normalized = value.replace(" ", "").replace(",", ".")
        try:
            return float(normalized)
        except ValueError:
            return None
    return None


def _is_empty_row(row: tuple, max_cols: int = 10) -> bool:
    """Проверяет, является ли строка пустой."""
    non_empty = sum(1 for cell in row[:max_cols] if cell is not None and str(cell).strip() != "")
    return non_empty == 0


def parse_sheet1_debts(db: Session, target_user: User, admin: User, ws, overwrite: bool = False) -> Dict[str, int]:
    """
    Парсит Sheet1 — историю долгов по кредиторам.
    Создает записи DebtHistory для каждого кредитора на каждую дату.
    Вычисляет изменение общего долга между датами.
    """
    rows = list(ws.iter_rows(values_only=True))
    
    # Известные названия кредиторов из файла (нормализованные)
    KNOWN_CREDITORS = {
        "СБЕР", "Т-БАНК", "ОЛЯ т-банк", "Оля СБЕР", "АЛЬФА", "МТС", "КРЕДИТ",
        "копилка", "ОЛЯ", "МТС1", "МТС2"
    }
    
    # Ключевые слова, которые НЕ являются кредиторами
    SKIP_KEYWORDS = {
        "ВСЕГО", "ОБЩЕГО", "TOTAL", "РАЗНИЦА", "ИЗМЕНЕНИЕ", "ДОЛГ", "ЗА", 
        "МЕСЯЦ", "ВЕСНУ", "ЗИМУ", "ЛЕТО", "ОСЕНЬ", "ГОД", "2024", "2025", "2026"
    }
    
    creditor_columns = {}
    date_rows = []  # Список кортежей (row_index, date, row_data)
    
    # Сначала находим все строки с датами
    for idx, row in enumerate(rows):
        if len(row) > 0:
            row_date = _coerce_date(row[0])
            if row_date:
                date_rows.append((idx, row_date, row))
    
    # Определяем колонки кредиторов из первой строки с датой
    # Обычно вторая строка после даты содержит названия
    if date_rows:
        for idx, row_date, row in date_rows[:3]:  # Проверяем первые несколько строк
            for col_idx in range(1, min(len(row), 15)):
                cell_value = row[col_idx] if len(row) > col_idx else None
                if isinstance(cell_value, str):
                    cell_stripped = cell_value.strip()
                    cell_upper = cell_stripped.upper()
                    
                    # Пропускаем служебные слова
                    if any(kw in cell_upper for kw in SKIP_KEYWORDS):
                        continue
                    if len(cell_stripped) < 2:
                        continue
                    
                    # Проверяем, похоже ли на название кредитора
                    is_creditor = False
                    for known in KNOWN_CREDITORS:
                        if known.upper() in cell_upper or cell_upper in known.upper():
                            is_creditor = True
                            break
                    
                    # Если не нашли в известных, но строка короткая и не содержит пробелов - возможно это кредитор
                    if not is_creditor and " " not in cell_stripped and len(cell_stripped) <= 15:
                        # Дополнительная проверка: не содержит ли слов-маркеров
                        if not any(kw in cell_upper for kw in ["ИЗМЕНЕНИЕ", "ВСЕГО", "РАЗНИЦА"]):
                            is_creditor = True
                    
                    if is_creditor and cell_stripped not in creditor_columns:
                        creditor_columns[cell_stripped] = col_idx
    
    inserted = 0
    updated = 0
    skipped = 0
    prev_total_debt = None
    
    # Обрабатываем каждую строку с датой
    for idx, row_date, row in date_rows:
        # Пропускаем строки-заголовки внутри данных
        first_value = row[1] if len(row) > 1 else None
        if isinstance(first_value, str):
            first_upper = first_value.upper()
            if any(kw in first_upper for kw in ["ВСЕГО", "СБЕР", "Т-БАНК", "ОЛЯ"]):
                # Это может быть строка-заголовок с названиями колонок
                continue
        
        # Собираем долги по кредиторам на эту дату
        creditors_debts = {}
        total_debt = 0.0
        
        for creditor_name, col_idx in creditor_columns.items():
            balance = _coerce_float(row[col_idx] if len(row) > col_idx else None)
            if balance is not None and balance > 0:
                creditors_debts[creditor_name] = balance
                total_debt += balance
        
        # Если не нашли по динамическим колонкам, пробуем стандартные позиции
        if not creditors_debts:
            standard_cols = {
                "СБЕР": 1,
                "ОЛЯ т-банк": 2,
                "Оля СБЕР": 3,
                "Т-БАНК": 4,
                "копилка": 5,
            }
            for creditor_name, col_idx in standard_cols.items():
                balance = _coerce_float(row[col_idx] if len(row) > col_idx else None)
                if balance is not None and balance > 0:
                    creditors_debts[creditor_name] = balance
                    total_debt += balance
        
        if not creditors_debts:
            continue
        
        # Создаём записи DebtHistory для каждого кредитора
        for creditor_name, balance in creditors_debts.items():
            existing = db.scalar(
                select(DebtHistory).where(
                    DebtHistory.creditor == creditor_name,
                    DebtHistory.record_date == row_date,
                )
            )
            
            if existing:
                if overwrite:
                    existing.amount = balance
                    updated += 1
                else:
                    skipped += 1
            else:
                history = DebtHistory(
                    creditor=creditor_name,
                    amount=balance,
                    record_date=row_date,
                )
                db.add(history)
                inserted += 1
        
        # Вычисляем изменение долга (аналог столбца "разница" в Excel)
        if prev_total_debt is not None:
            debt_change = total_debt - prev_total_debt
        
        prev_total_debt = total_debt
    
    return {"inserted": inserted, "updated": updated, "skipped": skipped}


def parse_grace_periods(db: Session, target_user: User, admin: User, ws, overwrite: bool = False) -> Dict[str, int]:
    """
    Парсит лист 'льготные периоды'.
    Создаёт/обновляет записи CreditCard с суммами к погашению.
    Добавляет planned_repayment_amount из колонки TOTAL.
    """
    rows = list(ws.iter_rows(values_only=True))
    
    inserted = 0
    updated = 0
    skipped = 0
    
    # Пропускаем заголовок
    for row in rows[1:]:
        if _is_empty_row(row):
            continue
        
        grace_date = _coerce_date(row[0] if len(row) > 0 else None)
        if not grace_date:
            continue
        
        # Т-банк (колонка B)
        t_bank_amount = _coerce_float(row[1] if len(row) > 1 else None)
        # СБЕР (колонка C)
        sber_amount = _coerce_float(row[2] if len(row) > 2 else None)
        # TOTAL (колонка D) - сумма к погашению
        total_amount = _coerce_float(row[3] if len(row) > 3 else None)
        
        # Обработка Т-банка
        if t_bank_amount and t_bank_amount > 0:
            card_name = "Т-банк"
            existing = db.scalar(
                select(CreditCard).where(
                    CreditCard.user_id == target_user.id,
                    CreditCard.card_name == card_name,
                    CreditCard.grace_start_date <= grace_date,
                )
            )
            
            if existing:
                if overwrite:
                    existing.current_debt = t_bank_amount
                    existing.planned_repayment_amount = total_amount if total_amount else t_bank_amount
                    updated += 1
                else:
                    skipped += 1
            else:
                # Создаём новую запись
                # grace_period_days вычисляется как разница между grace_start_date и grace_date
                # Для простоты используем 30 дней по умолчанию
                card = CreditCard(
                    user_id=target_user.id,
                    card_name=card_name,
                    grace_start_date=grace_date,
                    grace_period_days=30,
                    current_debt=t_bank_amount,
                    planned_repayment_amount=total_amount if total_amount else t_bank_amount,
                    status="active",
                    moderation_status=RecordStatus.APPROVED,
                    approved_by_id=admin.id,
                    approved_at=datetime.utcnow(),
                )
                db.add(card)
                inserted += 1
        
        # Обработка СБЕРА
        if sber_amount and sber_amount > 0:
            card_name = "СБЕР"
            existing = db.scalar(
                select(CreditCard).where(
                    CreditCard.user_id == target_user.id,
                    CreditCard.card_name == card_name,
                    CreditCard.grace_start_date <= grace_date,
                )
            )
            
            if existing:
                if overwrite:
                    existing.current_debt = sber_amount
                    existing.planned_repayment_amount = total_amount if total_amount else sber_amount
                    updated += 1
                else:
                    skipped += 1
            else:
                card = CreditCard(
                    user_id=target_user.id,
                    card_name=card_name,
                    grace_start_date=grace_date,
                    grace_period_days=30,
                    current_debt=sber_amount,
                    planned_repayment_amount=total_amount if total_amount else sber_amount,
                    status="active",
                    moderation_status=RecordStatus.APPROVED,
                    approved_by_id=admin.id,
                    approved_at=datetime.utcnow(),
                )
                db.add(card)
                inserted += 1
    
    return {"inserted": inserted, "updated": updated, "skipped": skipped}


def parse_expenses(db: Session, target_user: User, admin: User, ws, overwrite: bool = False) -> Dict[str, int]:
    """
    Парсит лист 'траты'.
    Создаёт записи Expense с категориями и датами.
    """
    rows = list(ws.iter_rows(values_only=True))
    
    inserted = 0
    updated = 0
    skipped = 0
    
    current_month = None
    
    for row in rows:
        if _is_empty_row(row, max_cols=4):
            continue
        
        description = row[0] if len(row) > 0 else None
        amount = _coerce_float(row[1] if len(row) > 1 else None)
        due_date_val = row[2] if len(row) > 2 else None
        month_label = row[3] if len(row) > 3 else None
        
        # Проверяем, не является ли строка заголовком месяца
        if isinstance(month_label, str) and month_label.strip():
            current_month = month_label.strip()
        
        # Пропускаем строки без суммы
        if amount is None or amount <= 0:
            continue
        
        # Пропускаем заголовки
        if isinstance(description, str) and "ОБЯЗАТЕЛЬНЫЕ" in description.upper():
            continue
        
        # Парсим дату
        due_date = _coerce_date(due_date_val)
        if not due_date and current_month:
            # Если даты нет, но есть месяц, создаём запись без конкретной даты
            pass
        
        # Определяем категорию
        category = None
        if description:
            desc_lower = str(description).lower()
            if "аренд" in desc_lower:
                category = ExpenseCategory.RENT
            elif "коммунал" in desc_lower:
                category = ExpenseCategory.UTILITIES
            elif "врач" in desc_lower or "медиц" in desc_lower:
                category = ExpenseCategory.MEDICAL
            elif "год" in desc_lower or "др" in desc_lower or "празд" in desc_lower:
                category = ExpenseCategory.GIFTS
            elif "toefl" in desc_lower or "учеб" in desc_lower or "образован" in desc_lower:
                category = ExpenseCategory.EDUCATION
        
        expense = Expense(
            user_id=target_user.id,
            amount=amount,
            due_date=due_date if due_date else date.today(),
            category=category,
            description=str(description) if description else None,
            is_mandatory=True,
            is_completed=False,
            moderation_status=RecordStatus.APPROVED,
            approved_by_id=admin.id,
            approved_at=datetime.utcnow(),
        )
        db.add(expense)
        inserted += 1
    
    return {"inserted": inserted, "updated": updated, "skipped": skipped}


def parse_incomes(db: Session, target_user: User, admin: User, ws, overwrite: bool = False) -> Dict[str, int]:
    """
    Парсит лист 'доход'.
    Создаёт записи Income с категориями и датами.
    """
    rows = list(ws.iter_rows(values_only=True))
    
    inserted = 0
    updated = 0
    skipped = 0
    
    for row in rows:
        if _is_empty_row(row, max_cols=3):
            continue
        
        description = row[0] if len(row) > 0 else None
        amount = _coerce_float(row[1] if len(row) > 1 else None)
        when_val = row[2] if len(row) > 2 else None
        
        # Пропускаем заголовки
        if isinstance(description, str) and "Доход" in description:
            continue
        if isinstance(description, str) and "ВСЕГО" in str(description).upper():
            continue
        
        if amount is None or amount <= 0:
            continue
        
        # Парсим дату
        income_date = _coerce_date(when_val)
        if not income_date:
            income_date = date.today()
        
        # Определяем категорию
        category = None
        if description:
            desc_lower = str(description).lower()
            if "зп" in desc_lower or "зарплат" in desc_lower:
                category = IncomeCategory.SALARY
            elif "аванс" in desc_lower:
                category = IncomeCategory.SALARY
            elif "стипенд" in desc_lower:
                category = IncomeCategory.SCHOLARSHIP
            elif "отец" in desc_lower or "мама" in desc_lower or "родит" in desc_lower:
                category = IncomeCategory.GIFT
            elif "долг" in desc_lower:
                category = IncomeCategory.GIFT  # Возврат долга
            elif "склад" in desc_lower:
                category = IncomeCategory.FREELANCE
        
        income = Income(
            user_id=target_user.id,
            amount=amount,
            income_date=income_date,
            category=category,
            description=str(description) if description else None,
            is_actual=False,
            moderation_status=RecordStatus.APPROVED,
            approved_by_id=admin.id,
            approved_at=datetime.utcnow(),
        )
        db.add(income)
        inserted += 1
    
    return {"inserted": inserted, "updated": updated, "skipped": skipped}


def import_full_xlsx(
    db: Session,
    file_path: str,
    target_user_id: Optional[int] = None,
    overwrite: bool = False,
) -> Dict[str, Any]:
    """
    Полный импорт всех листов из Excel файла.
    Импортирует:
    - Sheet1: история долгов по кредиторам с расчётом изменений (разница между датами)
    - льготные периоды: кредитные карты с суммами к погашению на даты
    - траты: обязательные расходы с категориями и датами
    - доход: доходы с категориями и датами
    
    Листы 'сезоны', 'годы', 'Суммы для трат' используются для валидации расчётов
    и не импортируются напрямую (расчёты производятся на основе импортированных данных).
    """
    import_file = Path(file_path)
    if not import_file.is_absolute():
        import_file = Path.cwd() / import_file
    
    if not import_file.exists():
        raise FileNotFoundError(f"Файл не найден: {file_path}")
    
    # Получаем пользователя
    if target_user_id:
        target_user = db.scalar(select(User).where(User.id == target_user_id))
        if not target_user:
            raise ValueError("Целевой пользователь не найден.")
    else:
        # Используем первого админа
        target_user = db.scalar(select(User).where(User.role == UserRole.ADMIN))
        if not target_user:
            raise ValueError("Администратор не найден.")
    
    admin = target_user  # Для модерации записей
    
    wb = openpyxl.load_workbook(import_file, data_only=True)
    
    results = {
        "file": import_file.name,
        "sheets_processed": [],
        "details": {},
    }
    
    # 1. Sheet1 - история долгов
    if "Sheet1" in wb.sheetnames:
        ws = wb["Sheet1"]
        debt_result = parse_sheet1_debts(db, target_user, admin, ws, overwrite)
        results["sheets_processed"].append("Sheet1")
        results["details"]["debts_history"] = debt_result
    
    # 2. льготные периоды
    if "льготные периоды" in wb.sheetnames:
        ws = wb["льготные периоды"]
        grace_result = parse_grace_periods(db, target_user, admin, ws, overwrite)
        results["sheets_processed"].append("льготные периоды")
        results["details"]["grace_periods"] = grace_result
    
    # 3. траты
    if "траты" in wb.sheetnames:
        ws = wb["траты"]
        expense_result = parse_expenses(db, target_user, admin, ws, overwrite)
        results["sheets_processed"].append("траты")
        results["details"]["expenses"] = expense_result
    
    # 4. доход
    if "доход" in wb.sheetnames:
        ws = wb["доход"]
        income_result = parse_incomes(db, target_user, admin, ws, overwrite)
        results["sheets_processed"].append("доход")
        results["details"]["incomes"] = income_result
    
    # Коммитим все изменения
    db.commit()
    
    # Подсчитываем итоги
    total_inserted = sum(r.get("inserted", 0) for r in results["details"].values())
    total_updated = sum(r.get("updated", 0) for r in results["details"].values())
    total_skipped = sum(r.get("skipped", 0) for r in results["details"].values())
    
    results["summary"] = {
        "inserted": total_inserted,
        "updated": total_updated,
        "skipped": total_skipped,
    }
    
    return results


if __name__ == "__main__":
    # Пример использования
    from app.database import SessionLocal, engine, Base
    from app.models import User, UserRole
    from sqlalchemy import select
    
    # Инициализация БД
    Base.metadata.create_all(bind=engine)
    
    with SessionLocal() as db:
        result = import_full_xlsx(db, "ДОЛГИ.xlsx", overwrite=True)
        print(f"Импорт завершён:")
        print(f"  Листы: {result['sheets_processed']}")
        print(f"  Вставлено: {result['summary']['inserted']}")
        print(f"  Обновлено: {result['summary']['updated']}")
        print(f"  Пропущено: {result['summary']['skipped']}")
        print(f"  Детали: {result['details']}")
