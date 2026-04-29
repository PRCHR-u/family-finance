import csv
from datetime import date, datetime, timedelta
from io import StringIO
from pathlib import Path
from typing import List, Dict, Any

import openpyxl
import pandas as pd
from fastapi import Depends, FastAPI, HTTPException, Query, UploadFile, File, status
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, Response
from fastapi.security import OAuth2PasswordBearer
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
from sqlalchemy import select, text
from sqlalchemy.orm import Session, aliased

from .auth import create_access_token, get_password_hash, verify_password, decode_access_token
from .database import Base, SessionLocal, engine, get_db
from .models import (
    AuditLog,
    CreditCard,
    CreditCardIssuer,
    CreditCardStatus,
    Creditor,
    Debt,
    DebtHistory,
    DebtRepayment,
    DebtStatus,
    Expense,
    ExpenseCategory,
    FinanceRecord,
    Income,
    IncomeCategory,
    RecordStatus,
    User,
    UserRole,
)
from .schemas import (
    AuditLogPage,
    AuditLogRead,
    BudgetSummaryResponse,
    ChangePasswordRequest,
    CreditCardCreate,
    CreditCardIssuerCreate,
    CreditCardIssuerRead,
    CreditCardIssuerUpdate,
    CreditCardRead,
    CreditorCreate,
    CreditorDebtHistory,
    CreditorRead,
    CreditorUpdate,
    DailyBudgetResponse,
    DebtChangeAnalysis,
    DebtCreate,
    DebtHistoryRead,
    DebtRead,
    DebtRepaymentCreate,
    DebtRepaymentRead,
    DebtSummary,
    ExpenseCreate,
    ExpenseRead,
    FinanceRecordCreate,
    FinanceRecordRead,
    FinanceRecordUpdate,
    IncomeCreate,
    IncomeRead,
    LoginRequest,
    TokenResponse,
    UserAdminUpdate,
    UserCreate,
    UserRead,
    WeeklyBudgetResponse,
)

app = FastAPI(title="Family Finance API")

# CORS middleware configuration
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:5173"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/auth/login")
STATIC_DIR = Path(__file__).resolve().parent.parent / "static"


class ImportResult(BaseModel):
    inserted: int
    updated: int
    skipped: int


class XLSXSheetData(BaseModel):
    """Модель для данных листа XLSX"""
    name: str
    columns: List[str]
    data: List[Dict[str, Any]]


class XLSXFileInfo(BaseModel):
    """Модель для информации о XLSX файле"""
    filename: str
    sheet_count: int
    sheets: List[str]


def _log_action(
    db: Session,
    action: str,
    actor_user_id: int | None = None,
    target_user_id: int | None = None,
    details: str | None = None,
):
    db.add(
        AuditLog(
            action=action,
            actor_user_id=actor_user_id,
            target_user_id=target_user_id,
            details=details,
        )
    )


def _ensure_users_table_columns():
    with engine.connect() as conn:
        existing_columns = {
            row[1] for row in conn.execute(text("PRAGMA table_info(users)")).fetchall()
        }
        if "hashed_password" not in existing_columns:
            conn.execute(
                text("ALTER TABLE users ADD COLUMN hashed_password VARCHAR(255) DEFAULT '' NOT NULL")
            )
            conn.commit()
        if "is_active" not in existing_columns:
            conn.execute(
                text("ALTER TABLE users ADD COLUMN is_active BOOLEAN DEFAULT 1 NOT NULL")
            )
            conn.commit()


@app.on_event("startup")
def on_startup():
    import os
    from .database import init_db
    
    # Инициализация БД через новую функцию с поддержкой переменных окружения
    init_db()
    
    # Дополнительная проверка наличия администратора (для обратной совместимости)
    _ensure_users_table_columns()
    with SessionLocal() as db:
        admin_username = os.getenv("ADMIN_USERNAME", "admin")
        admin_email = os.getenv("ADMIN_EMAIL", "admin@example.com")
        admin = db.scalar(select(User).where(User.username == admin_username))
        if not admin:
            # Пробуем найти любого админа
            admin = db.scalar(select(User).where(User.role == UserRole.ADMIN))
        if not admin:
            admin_password = os.getenv("ADMIN_PASSWORD", "ChangeMe123!")
            db.add(
                User(
                    username=admin_username,
                    role=UserRole.ADMIN,
                    hashed_password=get_password_hash(admin_password),
                    is_active=True
                )
            )
            db.commit()


def _coerce_date(value) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        try:
            return datetime.fromisoformat(value).date()
        except ValueError:
            return None
    return None


def _coerce_float(value) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        normalized = value.replace(" ", "").replace(",", ".")
        try:
            return float(normalized)
        except ValueError:
            return None
    return None


def _recalculate_debt_balance(db: Session, debt: Debt):
    approved_repayments = db.scalars(
        select(DebtRepayment).where(
            DebtRepayment.debt_id == debt.id,
            DebtRepayment.moderation_status == RecordStatus.APPROVED,
        )
    ).all()
    paid_amount = sum(item.amount for item in approved_repayments)
    debt.current_balance = max(0.0, debt.principal_amount - paid_amount)
    debt.status = DebtStatus.CLOSED if debt.current_balance <= 0 else DebtStatus.ACTIVE


def _debt_total_as_of(db: Session, current_user: User, as_of: date) -> float:
    stmt = select(Debt).where(
        Debt.moderation_status == RecordStatus.APPROVED,
        Debt.start_date <= as_of,
    )
    if current_user.role != UserRole.ADMIN:
        stmt = stmt.where(Debt.user_id == current_user.id)
    debts = db.scalars(stmt).all()
    total = 0.0
    for debt in debts:
        approved_paid = db.scalars(
            select(DebtRepayment).where(
                DebtRepayment.debt_id == debt.id,
                DebtRepayment.moderation_status == RecordStatus.APPROVED,
                DebtRepayment.payment_date <= as_of,
            )
        ).all()
        outstanding = max(0.0, debt.principal_amount - sum(item.amount for item in approved_paid))
        total += outstanding
    return total


def _serialize_credit_card(card: CreditCard) -> CreditCardRead:
    grace_end_date = card.grace_start_date + timedelta(days=card.grace_period_days)
    remaining_days = (grace_end_date - date.today()).days
    return CreditCardRead(
        id=card.id,
        user_id=card.user_id,
        card_name=card.card_name,
        grace_start_date=card.grace_start_date,
        grace_period_days=card.grace_period_days,
        current_debt=card.current_debt,
        status=card.status,
        moderation_status=card.moderation_status,
        approved_by_id=card.approved_by_id,
        approved_at=card.approved_at,
        created_at=card.created_at,
        updated_at=card.updated_at,
        comment=card.comment,
        grace_end_date=grace_end_date,
        remaining_grace_days=remaining_days,
        amount_to_pay_urgent=card.current_debt if card.status == CreditCardStatus.ACTIVE else 0.0,
    )


def get_current_user(token: str = Depends(oauth2_scheme), db: Session = Depends(get_db)) -> User:
    username = decode_access_token(token)
    if not username:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Недействительный токен.")
    user = db.scalar(select(User).where(User.username == username))
    if not user:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Пользователь не найден.")
    if not user.is_active:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Пользователь деактивирован.")
    return user


def require_admin(current_user: User = Depends(get_current_user)) -> User:
    if current_user.role != UserRole.ADMIN:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Нужны права администратора.")
    return current_user


@app.get("/health")
def health():
    return {"status": "ok"}


@app.post("/auth/register", response_model=UserRead, status_code=status.HTTP_201_CREATED)
def register(payload: UserCreate, db: Session = Depends(get_db)):
    existing = db.scalar(select(User).where(User.username == payload.username))
    if existing:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Пользователь уже существует.")
    user = User(
        username=payload.username,
        role=UserRole.USER,
        hashed_password=get_password_hash(payload.password),
    )
    db.add(user)
    db.flush()
    _log_action(db, "auth.register", actor_user_id=user.id, details=f"username={payload.username}")
    db.commit()
    db.refresh(user)
    return user


@app.post("/auth/login", response_model=TokenResponse)
def login(payload: LoginRequest, db: Session = Depends(get_db)):
    user = db.scalar(select(User).where(User.username == payload.username))
    if not user or not verify_password(payload.password, user.hashed_password):
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Неверный логин или пароль.")
    if not user.is_active:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Пользователь деактивирован.")
    return TokenResponse(access_token=create_access_token(subject=user.username))


@app.post("/users", response_model=UserRead, status_code=status.HTTP_201_CREATED)
def create_user(payload: UserCreate, db: Session = Depends(get_db), admin: User = Depends(require_admin)):
    existing = db.scalar(select(User).where(User.username == payload.username))
    if existing:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Пользователь уже существует.")
    user = User(
        username=payload.username,
        role=payload.role,
        hashed_password=get_password_hash(payload.password),
    )
    db.add(user)
    db.flush()
    _log_action(
        db,
        "users.create",
        actor_user_id=admin.id,
        target_user_id=user.id,
        details=f"username={user.username}, role={user.role.value}",
    )
    db.commit()
    db.refresh(user)
    return user


@app.get("/users", response_model=list[UserRead])
def list_users(db: Session = Depends(get_db), _: User = Depends(require_admin)):
    return db.scalars(select(User).order_by(User.id.asc())).all()


@app.patch("/users/{user_id}", response_model=UserRead)
def update_user_by_admin(
    user_id: int,
    payload: UserAdminUpdate,
    db: Session = Depends(get_db),
    admin: User = Depends(require_admin),
):
    user = db.scalar(select(User).where(User.id == user_id))
    if not user:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Пользователь не найден.")

    updates = payload.model_dump(exclude_unset=True)
    if not updates:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Нет данных для обновления.")

    if "username" in updates:
        existing = db.scalar(select(User).where(User.username == updates["username"], User.id != user.id))
        if existing:
            raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Логин уже занят.")
        user.username = updates["username"]
    if "role" in updates:
        user.role = updates["role"]
    if "password" in updates:
        user.hashed_password = get_password_hash(updates["password"])
    if "is_active" in updates:
        if user.id == admin.id and updates["is_active"] is False:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Нельзя деактивировать самого себя.")
        user.is_active = updates["is_active"]

    _log_action(
        db,
        "users.update",
        actor_user_id=admin.id,
        target_user_id=user.id,
        details=f"fields={','.join(updates.keys())}",
    )
    db.commit()
    db.refresh(user)
    return user


@app.delete("/users/{user_id}")
def delete_user_by_admin(user_id: int, db: Session = Depends(get_db), admin: User = Depends(require_admin)):
    user = db.scalar(select(User).where(User.id == user_id))
    if not user:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Пользователь не найден.")
    if user.id == admin.id:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Нельзя удалить самого себя.")
    has_records = db.scalar(
        select(FinanceRecord.id).where(FinanceRecord.user_id == user.id).limit(1)
    )
    if has_records:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Нельзя удалить пользователя с финансовыми записями. Сначала деактивируйте его.",
        )
    _log_action(db, "users.delete", actor_user_id=admin.id, target_user_id=user.id, details=user.username)
    db.delete(user)
    db.commit()
    return {"message": "Пользователь удален."}


@app.post("/users/{user_id}/deactivate", response_model=UserRead)
def deactivate_user(user_id: int, db: Session = Depends(get_db), admin: User = Depends(require_admin)):
    user = db.scalar(select(User).where(User.id == user_id))
    if not user:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Пользователь не найден.")
    if user.id == admin.id:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Нельзя деактивировать самого себя.")
    user.is_active = False
    _log_action(db, "users.deactivate", actor_user_id=admin.id, target_user_id=user.id)
    db.commit()
    db.refresh(user)
    return user


@app.post("/users/{user_id}/activate", response_model=UserRead)
def activate_user(user_id: int, db: Session = Depends(get_db), admin: User = Depends(require_admin)):
    user = db.scalar(select(User).where(User.id == user_id))
    if not user:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Пользователь не найден.")
    user.is_active = True
    _log_action(db, "users.activate", actor_user_id=admin.id, target_user_id=user.id)
    db.commit()
    db.refresh(user)
    return user


@app.get("/users/me", response_model=UserRead)
def get_me(current_user: User = Depends(get_current_user)):
    return current_user


@app.post("/users/change-password")
def change_password(
    payload: ChangePasswordRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if not verify_password(payload.old_password, current_user.hashed_password):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Старый пароль неверен.")
    current_user.hashed_password = get_password_hash(payload.new_password)
    _log_action(db, "users.change_password", actor_user_id=current_user.id)
    db.commit()
    return {"message": "Пароль обновлен."}


@app.post("/debts", response_model=DebtRead, status_code=status.HTTP_201_CREATED)
def create_debt(
    payload: DebtCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    moderation_status = RecordStatus.APPROVED if current_user.role == UserRole.ADMIN else RecordStatus.PENDING
    debt = Debt(
        user_id=current_user.id,
        current_balance=payload.principal_amount,
        moderation_status=moderation_status,
        approved_by_id=current_user.id if moderation_status == RecordStatus.APPROVED else None,
        approved_at=datetime.utcnow() if moderation_status == RecordStatus.APPROVED else None,
        **payload.model_dump(),
    )
    db.add(debt)
    db.flush()
    _log_action(
        db,
        "debts.create",
        actor_user_id=current_user.id,
        target_user_id=debt.user_id,
        details=f"debt_id={debt.id}, creditor={debt.creditor_name}",
    )
    db.commit()
    db.refresh(debt)
    return debt


@app.get("/debts", response_model=list[DebtRead])
def list_debts(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    user_id: int | None = Query(default=None),
    debt_status: DebtStatus | None = Query(default=None, alias="status"),
    moderation_status: RecordStatus | None = Query(default=None),
):
    stmt = select(Debt)
    if current_user.role != UserRole.ADMIN:
        stmt = stmt.where(Debt.user_id == current_user.id)
    elif user_id is not None:
        stmt = stmt.where(Debt.user_id == user_id)
    if debt_status is not None:
        stmt = stmt.where(Debt.status == debt_status)
    if moderation_status is not None:
        stmt = stmt.where(Debt.moderation_status == moderation_status)
    return db.scalars(stmt.order_by(Debt.start_date.desc(), Debt.id.desc())).all()


@app.post("/debts/{debt_id}/approve", response_model=DebtRead)
def approve_debt(debt_id: int, db: Session = Depends(get_db), admin: User = Depends(require_admin)):
    debt = db.scalar(select(Debt).where(Debt.id == debt_id))
    if not debt:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Долг не найден.")
    debt.moderation_status = RecordStatus.APPROVED
    debt.approved_by_id = admin.id
    debt.approved_at = datetime.utcnow()
    _log_action(
        db,
        "debts.approve",
        actor_user_id=admin.id,
        target_user_id=debt.user_id,
        details=f"debt_id={debt.id}",
    )
    db.commit()
    db.refresh(debt)
    return debt


@app.post("/debts/{debt_id}/reject", response_model=DebtRead)
def reject_debt(debt_id: int, db: Session = Depends(get_db), admin: User = Depends(require_admin)):
    debt = db.scalar(select(Debt).where(Debt.id == debt_id))
    if not debt:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Долг не найден.")
    debt.moderation_status = RecordStatus.REJECTED
    debt.approved_by_id = None
    debt.approved_at = None
    _log_action(
        db,
        "debts.reject",
        actor_user_id=admin.id,
        target_user_id=debt.user_id,
        details=f"debt_id={debt.id}",
    )
    db.commit()
    db.refresh(debt)
    return debt


@app.post("/debts/{debt_id}/repayments", response_model=DebtRepaymentRead, status_code=status.HTTP_201_CREATED)
def create_debt_repayment(
    debt_id: int,
    payload: DebtRepaymentCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    debt = db.scalar(select(Debt).where(Debt.id == debt_id))
    if not debt:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Долг не найден.")
    if current_user.role != UserRole.ADMIN and debt.user_id != current_user.id:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Нельзя добавлять платеж к чужому долгу.")

    moderation_status = RecordStatus.APPROVED if current_user.role == UserRole.ADMIN else RecordStatus.PENDING
    repayment = DebtRepayment(
        debt_id=debt.id,
        user_id=current_user.id,
        moderation_status=moderation_status,
        approved_by_id=current_user.id if moderation_status == RecordStatus.APPROVED else None,
        approved_at=datetime.utcnow() if moderation_status == RecordStatus.APPROVED else None,
        **payload.model_dump(),
    )
    db.add(repayment)
    if moderation_status == RecordStatus.APPROVED:
        _recalculate_debt_balance(db, debt)
    db.flush()
    _log_action(
        db,
        "debts.repayment.create",
        actor_user_id=current_user.id,
        target_user_id=debt.user_id,
        details=f"debt_id={debt.id}, repayment_id={repayment.id}, amount={repayment.amount}",
    )
    db.commit()
    db.refresh(repayment)
    return repayment


@app.get("/debts/{debt_id}/repayments", response_model=list[DebtRepaymentRead])
def list_debt_repayments(
    debt_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    debt = db.scalar(select(Debt).where(Debt.id == debt_id))
    if not debt:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Долг не найден.")
    if current_user.role != UserRole.ADMIN and debt.user_id != current_user.id:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Нельзя просматривать чужой долг.")
    stmt = select(DebtRepayment).where(DebtRepayment.debt_id == debt_id)
    return db.scalars(stmt.order_by(DebtRepayment.payment_date.asc(), DebtRepayment.id.asc())).all()


@app.post("/repayments/{repayment_id}/approve", response_model=DebtRepaymentRead)
def approve_repayment(
    repayment_id: int,
    db: Session = Depends(get_db),
    admin: User = Depends(require_admin),
):
    repayment = db.scalar(select(DebtRepayment).where(DebtRepayment.id == repayment_id))
    if not repayment:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Погашение не найдено.")
    debt = db.scalar(select(Debt).where(Debt.id == repayment.debt_id))
    if not debt:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Долг не найден.")
    repayment.moderation_status = RecordStatus.APPROVED
    repayment.approved_by_id = admin.id
    repayment.approved_at = datetime.utcnow()
    _recalculate_debt_balance(db, debt)
    _log_action(
        db,
        "debts.repayment.approve",
        actor_user_id=admin.id,
        target_user_id=debt.user_id,
        details=f"debt_id={debt.id}, repayment_id={repayment.id}",
    )
    db.commit()
    db.refresh(repayment)
    return repayment


@app.post("/repayments/{repayment_id}/reject", response_model=DebtRepaymentRead)
def reject_repayment(
    repayment_id: int,
    db: Session = Depends(get_db),
    admin: User = Depends(require_admin),
):
    repayment = db.scalar(select(DebtRepayment).where(DebtRepayment.id == repayment_id))
    if not repayment:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Погашение не найдено.")
    debt = db.scalar(select(Debt).where(Debt.id == repayment.debt_id))
    repayment.moderation_status = RecordStatus.REJECTED
    repayment.approved_by_id = None
    repayment.approved_at = None
    _log_action(
        db,
        "debts.repayment.reject",
        actor_user_id=admin.id,
        target_user_id=debt.user_id if debt else None,
        details=f"repayment_id={repayment.id}",
    )
    db.commit()
    db.refresh(repayment)
    return repayment


@app.post("/credit-cards", response_model=CreditCardRead, status_code=status.HTTP_201_CREATED)
def create_credit_card(
    payload: CreditCardCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    moderation_status = RecordStatus.APPROVED if current_user.role == UserRole.ADMIN else RecordStatus.PENDING
    card = CreditCard(
        user_id=current_user.id,
        status=CreditCardStatus.ACTIVE if payload.current_debt > 0 else CreditCardStatus.CLOSED,
        moderation_status=moderation_status,
        approved_by_id=current_user.id if moderation_status == RecordStatus.APPROVED else None,
        approved_at=datetime.utcnow() if moderation_status == RecordStatus.APPROVED else None,
        **payload.model_dump(),
    )
    db.add(card)
    db.flush()
    _log_action(
        db,
        "credit_cards.create",
        actor_user_id=current_user.id,
        target_user_id=card.user_id,
        details=f"card_id={card.id}, card_name={card.card_name}",
    )
    db.commit()
    db.refresh(card)
    return _serialize_credit_card(card)


@app.get("/credit-cards", response_model=list[CreditCardRead])
def list_credit_cards(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    user_id: int | None = Query(default=None),
    status_filter: CreditCardStatus | None = Query(default=None, alias="status"),
    moderation_status: RecordStatus | None = Query(default=None),
):
    stmt = select(CreditCard)
    if current_user.role != UserRole.ADMIN:
        stmt = stmt.where(CreditCard.user_id == current_user.id)
    elif user_id is not None:
        stmt = stmt.where(CreditCard.user_id == user_id)
    if status_filter is not None:
        stmt = stmt.where(CreditCard.status == status_filter)
    if moderation_status is not None:
        stmt = stmt.where(CreditCard.moderation_status == moderation_status)
    cards = db.scalars(stmt.order_by(CreditCard.grace_start_date.desc(), CreditCard.id.desc())).all()
    return [_serialize_credit_card(card) for card in cards]


@app.post("/credit-cards/{card_id}/approve", response_model=CreditCardRead)
def approve_credit_card(card_id: int, db: Session = Depends(get_db), admin: User = Depends(require_admin)):
    card = db.scalar(select(CreditCard).where(CreditCard.id == card_id))
    if not card:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Кредитная карта не найдена.")
    card.moderation_status = RecordStatus.APPROVED
    card.approved_by_id = admin.id
    card.approved_at = datetime.utcnow()
    _log_action(
        db,
        "credit_cards.approve",
        actor_user_id=admin.id,
        target_user_id=card.user_id,
        details=f"card_id={card.id}",
    )
    db.commit()
    db.refresh(card)
    return _serialize_credit_card(card)


@app.post("/credit-cards/{card_id}/reject", response_model=CreditCardRead)
def reject_credit_card(card_id: int, db: Session = Depends(get_db), admin: User = Depends(require_admin)):
    card = db.scalar(select(CreditCard).where(CreditCard.id == card_id))
    if not card:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Кредитная карта не найдена.")
    card.moderation_status = RecordStatus.REJECTED
    card.approved_by_id = None
    card.approved_at = None
    _log_action(
        db,
        "credit_cards.reject",
        actor_user_id=admin.id,
        target_user_id=card.user_id,
        details=f"card_id={card.id}",
    )
    db.commit()
    db.refresh(card)
    return _serialize_credit_card(card)


@app.get("/analytics/urgent-credit-cards", response_model=list[CreditCardRead])
def urgent_credit_cards(
    limit: int = Query(default=3, ge=1, le=20),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    stmt = select(CreditCard).where(
        CreditCard.moderation_status == RecordStatus.APPROVED,
        CreditCard.status == CreditCardStatus.ACTIVE,
        CreditCard.current_debt > 0,
    )
    if current_user.role != UserRole.ADMIN:
        stmt = stmt.where(CreditCard.user_id == current_user.id)
    cards = db.scalars(stmt).all()
    serialized = [_serialize_credit_card(card) for card in cards]
    serialized.sort(key=lambda item: (item.remaining_grace_days, -item.amount_to_pay_urgent))
    return serialized[:limit]


@app.post("/records", response_model=FinanceRecordRead, status_code=status.HTTP_201_CREATED)
def create_record(
    payload: FinanceRecordCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    initial_status = RecordStatus.APPROVED if current_user.role == UserRole.ADMIN else RecordStatus.PENDING
    record = FinanceRecord(
        user_id=current_user.id,
        status=initial_status,
        approved_by_id=current_user.id if initial_status == RecordStatus.APPROVED else None,
        approved_at=datetime.utcnow() if initial_status == RecordStatus.APPROVED else None,
        **payload.model_dump(),
    )
    db.add(record)
    db.commit()
    db.refresh(record)
    return record


@app.get("/records", response_model=list[FinanceRecordRead])
def list_records(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    status_filter: RecordStatus | None = Query(default=None, alias="status"),
    user_id: int | None = Query(default=None),
):
    stmt = select(FinanceRecord)
    if current_user.role != UserRole.ADMIN:
        stmt = stmt.where(FinanceRecord.user_id == current_user.id)
    elif user_id is not None:
        stmt = stmt.where(FinanceRecord.user_id == user_id)
    if status_filter is not None:
        stmt = stmt.where(FinanceRecord.status == status_filter)
    return db.scalars(stmt.order_by(FinanceRecord.period_date.asc(), FinanceRecord.id.asc())).all()


@app.put("/records/{record_id}", response_model=FinanceRecordRead)
def update_record(
    record_id: int,
    payload: FinanceRecordUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    record = db.scalar(select(FinanceRecord).where(FinanceRecord.id == record_id))
    if not record:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Запись не найдена.")
    if current_user.role != UserRole.ADMIN and record.user_id != current_user.id:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Нельзя менять чужую запись.")

    for key, value in payload.model_dump().items():
        setattr(record, key, value)

    if current_user.role == UserRole.ADMIN:
        record.status = RecordStatus.APPROVED
        record.approved_by_id = current_user.id
        record.approved_at = datetime.utcnow()
    else:
        record.status = RecordStatus.PENDING
        record.approved_by_id = None
        record.approved_at = None

    db.commit()
    db.refresh(record)
    return record


@app.post("/records/{record_id}/approve", response_model=FinanceRecordRead)
def approve_record(record_id: int, db: Session = Depends(get_db), admin: User = Depends(require_admin)):
    record = db.scalar(select(FinanceRecord).where(FinanceRecord.id == record_id))
    if not record:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Запись не найдена.")
    record.status = RecordStatus.APPROVED
    record.approved_by_id = admin.id
    record.approved_at = datetime.utcnow()
    _log_action(
        db,
        "records.approve",
        actor_user_id=admin.id,
        target_user_id=record.user_id,
        details=f"record_id={record.id}",
    )
    db.commit()
    db.refresh(record)
    return record


@app.post("/records/{record_id}/reject", response_model=FinanceRecordRead)
def reject_record(record_id: int, db: Session = Depends(get_db), admin: User = Depends(require_admin)):
    record = db.scalar(select(FinanceRecord).where(FinanceRecord.id == record_id))
    if not record:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Запись не найдена.")
    record.status = RecordStatus.REJECTED
    record.approved_by_id = None
    record.approved_at = None
    _log_action(
        db,
        "records.reject",
        actor_user_id=admin.id,
        target_user_id=record.user_id,
        details=f"record_id={record.id}",
    )
    db.commit()
    db.refresh(record)
    return record


@app.get("/analytics/debt-summary", response_model=DebtSummary)
def debt_summary(
    period: str = "month",
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    # Определяем диапазон дат на основе периода
    today = date.today()
    if period == "month":
        period_from = today.replace(day=1)
        period_to = today
    elif period == "year":
        period_from = today.replace(month=1, day=1)
        period_to = today
    else:
        # По умолчанию берем текущий месяц
        period_from = today.replace(day=1)
        period_to = today

    if period_from > period_to:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Неверный диапазон дат.")

    day_before = period_from.fromordinal(period_from.toordinal() - 1)
    opening_debt = _debt_total_as_of(db, current_user, day_before)
    closing_debt = _debt_total_as_of(db, current_user, period_to)

    finance_stmt = select(FinanceRecord).where(
        FinanceRecord.status == RecordStatus.APPROVED,
        FinanceRecord.period_date >= period_from,
        FinanceRecord.period_date <= period_to,
    )
    if current_user.role != UserRole.ADMIN:
        finance_stmt = finance_stmt.where(FinanceRecord.user_id == current_user.id)
    records = db.scalars(
        finance_stmt.order_by(FinanceRecord.period_date.asc(), FinanceRecord.id.asc())
    ).all()

    if not records:
        return DebtSummary(
            period_from=period_from,
            period_to=period_to,
            records_count=0,
            opening_debt=opening_debt,
            closing_debt=closing_debt,
            debt_change=closing_debt - opening_debt,
            total_income=0.0,
            total_mandatory_expense=0.0,
            total_urgent_creditcard_repayment=0.0,
        )

    return DebtSummary(
        period_from=period_from,
        period_to=period_to,
        records_count=len(records),
        opening_debt=opening_debt,
        closing_debt=closing_debt,
        debt_change=closing_debt - opening_debt,
        total_income=sum(item.income for item in records),
        total_mandatory_expense=sum(item.mandatory_expense for item in records),
        total_urgent_creditcard_repayment=sum(item.urgent_creditcard_repayment for item in records),
    )


# ==================== BUDGET ANALYSIS ENDPOINTS ====================

@app.get("/analytics/debt-change", response_model=DebtChangeAnalysis)
def analyze_debt_change(
    period_from: date,
    period_to: date,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Анализирует прирост и уменьшение долга за указанный период.
    
    Возвращает детальную информацию:
    - Долг на начало и конец периода
    - Новые долги, созданные в периоде
    - Платежи по долгам в периоде
    - Чистое изменение долга
    """
    from .budget_utils import get_debt_change_analysis
    
    if period_from > period_to:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Неверный диапазон дат.")
    
    result = get_debt_change_analysis(db, current_user, period_from, period_to)
    return result


@app.get("/analytics/weekly-budget", response_model=WeeklyBudgetResponse)
def get_weekly_budget(
    reference_date: date | None = Query(default=None),
    weeks_ahead: int = Query(default=1, ge=1, le=12),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Рассчитывает недельный бюджет на основе анализа расходов.
    
    Параметры:
    - reference_date: дата начала периода (по умолчанию сегодня)
    - weeks_ahead: количество недель для планирования (1-12)
    
    Возвращает:
    - Общий бюджет на неделю
    - Средний дневной бюджет
    - Обязательные расходы
    - Доходы за период
    - Рекомендации по бюджету
    """
    from .budget_utils import calculate_weekly_budget
    
    result = calculate_weekly_budget(db, current_user, reference_date, weeks_ahead)
    return result


@app.get("/analytics/daily-budget", response_model=DailyBudgetResponse)
def get_daily_budget(
    target_date: date | None = Query(default=None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Рассчитывает ежедневный бюджет на конкретный день (по запросу).
    
    Параметры:
    - target_date: дата расчета (по умолчанию сегодня)
    
    Возвращает:
    - Обязательные расходы на день
    - Дискреционный бюджет (на необязательные траты)
    - Рекомендации
    """
    from .budget_utils import calculate_daily_budget
    
    result = calculate_daily_budget(db, current_user, target_date)
    return result


@app.get("/analytics/budget-summary", response_model=BudgetSummaryResponse)
def get_budget_summary_endpoint(
    period_from: date,
    period_to: date,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Сводная информация о бюджете за период с анализом долгов.
    
    Объединяет анализ изменения долгов и бюджетирование.
    """
    from .budget_utils import get_budget_summary
    
    if period_from > period_to:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Неверный диапазон дат.")
    
    result = get_budget_summary(db, current_user, period_from, period_to)
    return result


@app.get("/analytics/debt-timeline", response_model=list)
def get_debt_timeline(
    start_date: date,
    end_date: date,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    История изменения долгов по датам (как в таблице Excel Sheet1).
    
    Возвращает список записей с полями:
    - date: дата
    - creditors: словарь {кредитор: сумма}
    - total_debt: общий долг на дату
    - debt_change: изменение долга с предыдущей даты
    """
    from .budget_utils import get_debt_history_timeline
    
    if start_date > end_date:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Неверный диапазон дат.")
    
    return get_debt_history_timeline(db, current_user, start_date, end_date)


@app.get("/analytics/seasonal-debt", response_model=dict)
def get_seasonal_debt(
    year: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Агрегированные показатели долга по сезонам (аналог листа 'сезоны' в Excel).
    
    Возвращает данные по сезонам: зима, весна, лето, осень.
    Для каждого сезона: opening_debt, closing_debt, debt_change.
    """
    from .budget_utils import get_seasonal_debt_summary
    
    return get_seasonal_debt_summary(db, current_user, year)


@app.get("/analytics/yearly-debt", response_model=dict)
def get_yearly_debt(
    year: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Агрегированные показатели долга за год (аналог листа 'годы' в Excel).
    
    Возвращает: opening_debt, closing_debt, debt_change за указанный год.
    """
    from .budget_utils import get_yearly_debt_summary
    
    return get_yearly_debt_summary(db, current_user, year)


@app.post("/imports/excel", response_model=ImportResult)
def import_excel(
    file_path: str = "ДОЛГИ.xlsx",
    overwrite: bool = False,
    target_user_id: int | None = None,
    db: Session = Depends(get_db),
    admin: User = Depends(require_admin),
):
    """
    Импорт детализированных данных из Excel файла ДОЛГИ.xlsx:
    - Создаёт объекты Debt для каждого кредитора (СБЕР, АЛЬФА, МТС1, МТС2, Т-БАНК, ОЛЯ, КРЕДИТ)
    - Создаёт объекты Income из блока доходов (колонки M-O)
    - Создаёт объекты Expense из блока обязательных трат (колонки M-O)
    """
    import_file = Path(file_path)
    if not import_file.is_absolute():
        import_file = Path.cwd() / import_file
    if not import_file.exists():
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Файл не найден.")

    target_user = admin if target_user_id is None else db.scalar(select(User).where(User.id == target_user_id))
    if not target_user:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Целевой пользователь не найден.")

    wb = openpyxl.load_workbook(import_file, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))

    # Словарь кредиторов и их колонок
    CREDITOR_COLUMNS = {
        "СБЕР": 1,      # B
        "АЛЬФА": 2,     # C
        "МТС1": 3,      # D
        "МТС2": 4,      # E
        "Т-БАНК": 5,    # F
        "ОЛЯ": 6,       # G
        "КРЕДИТ": 7,    # H
    }

    inserted_debts = 0
    inserted_incomes = 0
    inserted_expenses = 0
    updated = 0
    skipped = 0

    # Парсинг строк с долгами (колонки A-I)
    for row in rows:
        period_date = _coerce_date(row[0] if len(row) > 0 else None)
        if not period_date or not isinstance(period_date, date):
            continue

        # Создаём записи долгов по каждому кредитору
        for creditor_name, col_idx in CREDITOR_COLUMNS.items():
            balance = _coerce_float(row[col_idx] if len(row) > col_idx else None)
            if balance is None or balance <= 0:
                continue

            # Проверяем существующий долг
            existing_debt = db.scalar(
                select(Debt).where(
                    Debt.user_id == target_user.id,
                    Debt.creditor_name == creditor_name,
                    Debt.start_date == period_date,
                )
            )

            if existing_debt and not overwrite:
                skipped += 1
                continue

            if existing_debt:
                existing_debt.current_balance = balance
                existing_debt.principal_amount = max(existing_debt.principal_amount, balance)
                updated += 1
            else:
                debt = Debt(
                    user_id=target_user.id,
                    creditor_name=creditor_name,
                    principal_amount=balance,
                    current_balance=balance,
                    start_date=period_date,
                    moderation_status=RecordStatus.APPROVED,
                    approved_by_id=admin.id,
                    approved_at=datetime.utcnow(),
                )
                db.add(debt)
                inserted_debts += 1

    # Парсинг доходов и расходов (колонки M-O)
    current_section = None  # 'income' или 'expense'

    for row in rows:
        if len(row) <= 12:
            continue

        cell_m = row[12]  # Название/категория
        cell_n = row[13]  # Сумма
        cell_o = row[14]  # Дата

        # Определяем секцию по заголовкам
        if isinstance(cell_m, str):
            if "доход" in cell_m.lower():
                current_section = "income"
                continue
            elif "трат" in cell_m.lower() or "расход" in cell_m.lower():
                current_section = "expense"
                continue
            elif "ВСЕГО" in cell_m or "TOTAL" in cell_m or "изменение" in cell_m.lower():
                current_section = None
                continue

        if current_section is None:
            continue

        # Пропускаем строки-заголовки внутри секций
        if isinstance(cell_n, str) or cell_n is None:
            continue

        amount = _coerce_float(cell_n)
        if amount is None or amount <= 0:
            continue

        # Парсим дату
        item_date = _coerce_date(cell_o)
        if item_date is None:
            # Попытка распарсить строку вида "17.04-20.04"
            if isinstance(cell_o, str):
                try:
                    # Берём первую дату из диапазона
                    date_part = cell_o.split("-")[0].strip()
                    item_date = datetime.strptime(date_part, "%d.%m").replace(year=2026).date()
                except (ValueError, IndexError):
                    continue
            else:
                continue

        description = cell_m if isinstance(cell_m, str) else None

        if current_section == "income":
            # Определяем категорию дохода
            category = None
            if description:
                desc_lower = description.lower()
                if "зп" in desc_lower or "зарплат" in desc_lower:
                    category = IncomeCategory.SALARY
                elif "аванс" in desc_lower:
                    category = IncomeCategory.SALARY
                elif "стипенд" in desc_lower:
                    category = IncomeCategory.SCHOLARSHIP
                elif "отец" in desc_lower or "мама" in desc_lower or "папа" in desc_lower:
                    category = IncomeCategory.GIFT
                elif "склад" in desc_lower:
                    category = IncomeCategory.FREELANCE

            income = Income(
                user_id=target_user.id,
                amount=amount,
                income_date=item_date,
                category=category,
                description=description,
                is_actual=False,  # По умолчанию планируемый
                moderation_status=RecordStatus.APPROVED,
                approved_by_id=admin.id,
                approved_at=datetime.utcnow(),
            )
            db.add(income)
            inserted_incomes += 1

        elif current_section == "expense":
            # Определяем категорию расхода
            category = None
            is_mandatory = True  # Все из этого блока считаем обязательными
            if description:
                desc_lower = description.lower()
                if "аренд" in desc_lower:
                    category = ExpenseCategory.RENT
                elif "коммунал" in desc_lower:
                    category = ExpenseCategory.UTILITIES
                elif "белка" in desc_lower or "продукт" in desc_lower or "еда" in desc_lower:
                    category = ExpenseCategory.FOOD
                elif "топлив" in desc_lower or "бензин" in desc_lower:
                    category = ExpenseCategory.TRANSPORT
                elif "год" in desc_lower or "др" in desc_lower or "подар" in desc_lower:
                    category = ExpenseCategory.GIFTS
                elif "toefl" in desc_lower or "учеб" in desc_lower:
                    category = ExpenseCategory.EDUCATION

            expense = Expense(
                user_id=target_user.id,
                amount=amount,
                due_date=item_date,
                category=category,
                description=description,
                is_mandatory=is_mandatory,
                is_completed=False,
                moderation_status=RecordStatus.APPROVED,
                approved_by_id=admin.id,
                approved_at=datetime.utcnow(),
            )
            db.add(expense)
            inserted_expenses += 1

    _log_action(
        db,
        "imports.excel",
        actor_user_id=admin.id,
        target_user_id=target_user.id,
        details=f"file={import_file.name}, debts={inserted_debts}, incomes={inserted_incomes}, expenses={inserted_expenses}, updated={updated}, skipped={skipped}",
    )
    db.commit()
    return ImportResult(inserted=inserted_debts + inserted_incomes + inserted_expenses, updated=updated, skipped=skipped)


@app.post("/imports/excel-full", response_model=Dict[str, Any])
def import_excel_full(
    file_path: str = "ДОЛГИ.xlsx",
    overwrite: bool = False,
    target_user_id: int | None = None,
    db: Session = Depends(get_db),
    admin: User = Depends(require_admin),
):
    """
    Расширенный импорт всех листов из Excel файла ДОЛГИ.xlsx:
    - Sheet1: история долгов по кредиторам с расчётом изменений (разница между датами)
    - льготные периоды: кредитные карты с суммами к погашению на даты
    - траты: обязательные расходы с категориями и датами
    - доход: доходы с категориями и датами
    
    Воспроизводит логику расчётов "как в таблице".
    """
    from .xlsx_full_import import import_full_xlsx
    
    try:
        result = import_full_xlsx(
            db=db,
            file_path=file_path,
            target_user_id=target_user_id,
            overwrite=overwrite,
        )
        
        _log_action(
            db,
            "imports.excel_full",
            actor_user_id=admin.id,
            target_user_id=target_user_id or admin.id,
            details=f"file={file_path}, sheets={result['sheets_processed']}, inserted={result['summary']['inserted']}, updated={result['summary']['updated']}, skipped={result['summary']['skipped']}",
        )
        
        return result
    except FileNotFoundError as e:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=str(e))
    except ValueError as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"Ошибка импорта: {str(e)}")


# ==================== INCOME ENDPOINTS ====================

@app.post("/incomes", response_model=IncomeRead, status_code=status.HTTP_201_CREATED)
def create_income(
    payload: IncomeCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    moderation_status = RecordStatus.APPROVED if current_user.role == UserRole.ADMIN else RecordStatus.PENDING
    income = Income(
        user_id=current_user.id,
        moderation_status=moderation_status,
        approved_by_id=current_user.id if moderation_status == RecordStatus.APPROVED else None,
        approved_at=datetime.utcnow() if moderation_status == RecordStatus.APPROVED else None,
        **payload.model_dump(),
    )
    db.add(income)
    db.flush()
    _log_action(
        db,
        "incomes.create",
        actor_user_id=current_user.id,
        target_user_id=income.user_id,
        details=f"income_id={income.id}, amount={income.amount}",
    )
    db.commit()
    db.refresh(income)
    return income


@app.get("/incomes", response_model=list[IncomeRead])
def list_incomes(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    user_id: int | None = Query(default=None),
    moderation_status: RecordStatus | None = Query(default=None),
    is_actual: bool | None = Query(default=None),
):
    stmt = select(Income)
    if current_user.role != UserRole.ADMIN:
        stmt = stmt.where(Income.user_id == current_user.id)
    elif user_id is not None:
        stmt = stmt.where(Income.user_id == user_id)
    if moderation_status is not None:
        stmt = stmt.where(Income.moderation_status == moderation_status)
    if is_actual is not None:
        stmt = stmt.where(Income.is_actual == is_actual)
    return db.scalars(stmt.order_by(Income.income_date.desc(), Income.id.desc())).all()


@app.post("/incomes/{income_id}/approve", response_model=IncomeRead)
def approve_income(income_id: int, db: Session = Depends(get_db), admin: User = Depends(require_admin)):
    income = db.scalar(select(Income).where(Income.id == income_id))
    if not income:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Доход не найден.")
    income.moderation_status = RecordStatus.APPROVED
    income.approved_by_id = admin.id
    income.approved_at = datetime.utcnow()
    _log_action(
        db,
        "incomes.approve",
        actor_user_id=admin.id,
        target_user_id=income.user_id,
        details=f"income_id={income.id}",
    )
    db.commit()
    db.refresh(income)
    return income


@app.post("/incomes/{income_id}/mark-actual", response_model=IncomeRead)
def mark_income_actual(income_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    income = db.scalar(select(Income).where(Income.id == income_id))
    if not income:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Доход не найден.")
    if current_user.role != UserRole.ADMIN and income.user_id != current_user.id:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Нет доступа.")
    income.is_actual = True
    db.commit()
    db.refresh(income)
    return income


# ==================== EXPENSE ENDPOINTS ====================

@app.post("/expenses", response_model=ExpenseRead, status_code=status.HTTP_201_CREATED)
def create_expense(
    payload: ExpenseCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    moderation_status = RecordStatus.APPROVED if current_user.role == UserRole.ADMIN else RecordStatus.PENDING
    expense = Expense(
        user_id=current_user.id,
        moderation_status=moderation_status,
        approved_by_id=current_user.id if moderation_status == RecordStatus.APPROVED else None,
        approved_at=datetime.utcnow() if moderation_status == RecordStatus.APPROVED else None,
        **payload.model_dump(),
    )
    db.add(expense)
    db.flush()
    _log_action(
        db,
        "expenses.create",
        actor_user_id=current_user.id,
        target_user_id=expense.user_id,
        details=f"expense_id={expense.id}, amount={expense.amount}",
    )
    db.commit()
    db.refresh(expense)
    return expense


@app.get("/expenses", response_model=list[ExpenseRead])
def list_expenses(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    user_id: int | None = Query(default=None),
    moderation_status: RecordStatus | None = Query(default=None),
    is_completed: bool | None = Query(default=None),
    is_mandatory: bool | None = Query(default=None),
):
    stmt = select(Expense)
    if current_user.role != UserRole.ADMIN:
        stmt = stmt.where(Expense.user_id == current_user.id)
    elif user_id is not None:
        stmt = stmt.where(Expense.user_id == user_id)
    if moderation_status is not None:
        stmt = stmt.where(Expense.moderation_status == moderation_status)
    if is_completed is not None:
        stmt = stmt.where(Expense.is_completed == is_completed)
    if is_mandatory is not None:
        stmt = stmt.where(Expense.is_mandatory == is_mandatory)
    return db.scalars(stmt.order_by(Expense.due_date.asc(), Expense.id.desc())).all()


@app.post("/expenses/{expense_id}/approve", response_model=ExpenseRead)
def approve_expense(expense_id: int, db: Session = Depends(get_db), admin: User = Depends(require_admin)):
    expense = db.scalar(select(Expense).where(Expense.id == expense_id))
    if not expense:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Расход не найден.")
    expense.moderation_status = RecordStatus.APPROVED
    expense.approved_by_id = admin.id
    expense.approved_at = datetime.utcnow()
    _log_action(
        db,
        "expenses.approve",
        actor_user_id=admin.id,
        target_user_id=expense.user_id,
        details=f"expense_id={expense.id}, is_completed={expense.is_completed}",
    )
    db.commit()
    db.refresh(expense)
    return expense


@app.post("/expenses/{expense_id}/reject", response_model=ExpenseRead)
def reject_expense(expense_id: int, db: Session = Depends(get_db), admin: User = Depends(require_admin)):
    """Отклонить изменение статуса расхода."""
    expense = db.scalar(select(Expense).where(Expense.id == expense_id))
    if not expense:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Расход не найден.")
    expense.moderation_status = RecordStatus.REJECTED
    expense.approved_by_id = admin.id
    expense.approved_at = datetime.utcnow()
    # Откат изменения is_completed при отклонении
    expense.is_completed = False
    _log_action(
        db,
        "expenses.reject",
        actor_user_id=admin.id,
        target_user_id=expense.user_id,
        details=f"expense_id={expense.id}",
    )
    db.commit()
    db.refresh(expense)
    return expense


@app.post("/expenses/{expense_id}/complete", response_model=ExpenseRead)
def complete_expense(expense_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    """Отметить расход как выполненный (требует подтверждения администратором)."""
    expense = db.scalar(select(Expense).where(Expense.id == expense_id))
    if not expense:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Расход не найден.")
    if current_user.role != UserRole.ADMIN and expense.user_id != current_user.id:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Нет доступа.")
    
    # Изменение статуса требует модерации
    expense.is_completed = True
    expense.moderation_status = RecordStatus.PENDING
    expense.approved_by_id = None
    expense.approved_at = None
    
    _log_action(
        db,
        "expenses.complete",
        actor_user_id=current_user.id,
        target_user_id=expense.user_id,
        details=f"expense_id={expense_id}",
    )
    db.commit()
    db.refresh(expense)
    return expense


@app.get("/analytics/mandatory-expenses-monthly", response_model=dict)
def get_monthly_mandatory_expenses(
    year: int = Query(default=None),
    month: int = Query(default=None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    from datetime import date as dt_date
    
    today = dt_date.today()
    if year is None:
        year = today.year
    if month is None:
        month = today.month
    
    from calendar import monthrange
    _, last_day = monthrange(year, month)
    start_date = dt_date(year, month, 1)
    end_date = dt_date(year, month, last_day)
    
    stmt = select(Expense).where(
        Expense.user_id == current_user.id if current_user.role != UserRole.ADMIN else True,
        Expense.is_mandatory == True,
        Expense.due_date >= start_date,
        Expense.due_date <= end_date,
    )
    expenses = db.scalars(stmt).all()
    
    total = sum(e.amount for e in expenses)
    completed = sum(e.amount for e in expenses if e.is_completed)
    
    return {
        "year": year,
        "month": month,
        "total_mandatory": total,
        "completed": completed,
        "pending": total - completed,
        "count": len(expenses),
    }


# ==================== AUDIT LOGS ====================

def _build_audit_query(
    action: str | None,
    date_from: date | None,
    date_to: date | None,
    actor_user_id: int | None,
    target_user_id: int | None,
):
    actor_user = aliased(User)
    target_user = aliased(User)
    stmt = (
        select(
            AuditLog,
            actor_user.username.label("actor_username"),
            target_user.username.label("target_username"),
        )
        .outerjoin(actor_user, AuditLog.actor_user_id == actor_user.id)
        .outerjoin(target_user, AuditLog.target_user_id == target_user.id)
    )
    if action:
        stmt = stmt.where(AuditLog.action == action)
    if date_from:
        stmt = stmt.where(AuditLog.created_at >= datetime.combine(date_from, datetime.min.time()))
    if date_to:
        stmt = stmt.where(AuditLog.created_at <= datetime.combine(date_to, datetime.max.time()))
    if actor_user_id is not None:
        stmt = stmt.where(AuditLog.actor_user_id == actor_user_id)
    if target_user_id is not None:
        stmt = stmt.where(AuditLog.target_user_id == target_user_id)
    stmt = stmt.order_by(AuditLog.id)
    return stmt


def _map_audit_rows(result_rows) -> list[AuditLogRead]:
    return [
        AuditLogRead(
            id=row[0].id,
            action=row[0].action,
            actor_user_id=row[0].actor_user_id,
            actor_username=row[1],
            target_user_id=row[0].target_user_id,
            target_username=row[2],
            details=row[0].details,
            created_at=row[0].created_at,
        )
        for row in result_rows
    ]


@app.get("/audit-logs", response_model=AuditLogPage)
def list_audit_logs(
    action: str | None = Query(default=None),
    date_from: date | None = Query(default=None),
    date_to: date | None = Query(default=None),
    actor_user_id: int | None = Query(default=None),
    target_user_id: int | None = Query(default=None),
    limit: int = Query(default=100, ge=1, le=500),
    offset: int = Query(default=0, ge=0),
    db: Session = Depends(get_db),
    _: User = Depends(require_admin),
):
    base_stmt = _build_audit_query(action, date_from, date_to, actor_user_id, target_user_id)
    rows = db.execute(base_stmt).all()
    total_count = len(rows)
    paged_rows = rows[offset : offset + limit]
    return AuditLogPage(
        items=_map_audit_rows(paged_rows),
        total_count=total_count,
        limit=limit,
        offset=offset,
    )


@app.get("/audit-logs/export.csv")
def export_audit_logs_csv(
    action: str | None = Query(default=None),
    date_from: date | None = Query(default=None),
    date_to: date | None = Query(default=None),
    actor_user_id: int | None = Query(default=None),
    target_user_id: int | None = Query(default=None),
    db: Session = Depends(get_db),
    _: User = Depends(require_admin),
):
    stmt = _build_audit_query(action, date_from, date_to, actor_user_id, target_user_id)
    stmt = stmt.order_by(AuditLog.created_at.desc(), AuditLog.id.desc())
    logs = _map_audit_rows(db.execute(stmt).all())

    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(
        ["created_at", "action", "actor_user_id", "actor_username", "target_user_id", "target_username", "details"]
    )
    for item in logs:
        writer.writerow(
            [
                item.created_at.isoformat(),
                item.action,
                item.actor_user_id,
                item.actor_username,
                item.target_user_id,
                item.target_username,
                item.details or "",
            ]
        )

    return Response(
        content=output.getvalue(),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": 'attachment; filename="audit_logs.csv"'},
    )


app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")


@app.get("/")
def root():
    from fastapi.responses import FileResponse
    return FileResponse(STATIC_DIR / "index.html")


@app.get("/assets/{path:path}")
def serve_assets(path: str):
    """Serve frontend assets (JS, CSS) from static directory."""
    from fastapi.responses import FileResponse
    file_path = STATIC_DIR / "assets" / path
    if file_path.exists():
        return FileResponse(file_path)
    raise HTTPException(status_code=404, detail="Asset not found")


@app.get("/{path:path}")
def serve_static_files(path: str):
    """Serve static files (favicon, icons, etc.) from static directory."""
    from fastapi.responses import FileResponse
    file_path = STATIC_DIR / path
    if file_path.exists() and file_path.is_file():
        return FileResponse(file_path)
    raise HTTPException(status_code=404, detail="File not found")


@app.get("/assets/{path:path}")
def serve_assets(path: str):
    """Serve frontend assets (JS, CSS) from static directory."""
    from fastapi.responses import FileResponse
    file_path = STATIC_DIR / "assets" / path
    if file_path.exists():
        return FileResponse(file_path)
    raise HTTPException(status_code=404, detail="Asset not found")


# ==================== DICTIONARY ENDPOINTS (CREDITORS) ====================

@app.post("/creditors", response_model=CreditorRead, status_code=status.HTTP_201_CREATED)
def create_creditor(
    payload: CreditorCreate,
    db: Session = Depends(get_db),
    admin: User = Depends(require_admin),
):
    """Создать запись в справочнике кредиторов."""
    existing = db.scalar(select(Creditor).where(Creditor.name == payload.name))
    if existing:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Кредитор с таким именем уже существует.")
    creditor = Creditor(**payload.model_dump())
    db.add(creditor)
    db.flush()
    _log_action(
        db,
        "creditors.create",
        actor_user_id=admin.id,
        target_user_id=None,
        details=f"creditor_id={creditor.id}, name={creditor.name}",
    )
    db.commit()
    db.refresh(creditor)
    return creditor


@app.get("/creditors", response_model=list[CreditorRead])
def list_creditors(
    db: Session = Depends(get_db),
    is_active: bool | None = Query(default=None),
):
    """Получить список кредиторов."""
    stmt = select(Creditor)
    if is_active is not None:
        stmt = stmt.where(Creditor.is_active == is_active)
    return db.scalars(stmt.order_by(Creditor.name.asc())).all()


@app.patch("/creditors/{creditor_id}", response_model=CreditorRead)
def update_creditor(
    creditor_id: int,
    payload: CreditorUpdate,
    db: Session = Depends(get_db),
    admin: User = Depends(require_admin),
):
    """Обновить запись в справочнике кредиторов."""
    creditor = db.scalar(select(Creditor).where(Creditor.id == creditor_id))
    if not creditor:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Кредитор не найден.")
    
    updates = payload.model_dump(exclude_unset=True)
    if not updates:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Нет данных для обновления.")
    
    if "name" in updates:
        existing = db.scalar(select(Creditor).where(Creditor.name == updates["name"], Creditor.id != creditor.id))
        if existing:
            raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Кредитор с таким именем уже существует.")
        creditor.name = updates["name"]
    if "description" in updates:
        creditor.description = updates["description"]
    if "is_active" in updates:
        creditor.is_active = updates["is_active"]
    
    _log_action(
        db,
        "creditors.update",
        actor_user_id=admin.id,
        target_user_id=None,
        details=f"creditor_id={creditor.id}, fields={','.join(updates.keys())}",
    )
    db.commit()
    db.refresh(creditor)
    return creditor


@app.delete("/creditors/{creditor_id}")
def delete_creditor(creditor_id: int, db: Session = Depends(get_db), admin: User = Depends(require_admin)):
    """Удалить запись из справочника кредиторов."""
    creditor = db.scalar(select(Creditor).where(Creditor.id == creditor_id))
    if not creditor:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Кредитор не найден.")
    
    _log_action(
        db,
        "creditors.delete",
        actor_user_id=admin.id,
        target_user_id=None,
        details=f"creditor_id={creditor_id}, name={creditor.name}",
    )
    db.delete(creditor)
    db.commit()
    return {"message": "Кредитор удалён."}


# ==================== DICTIONARY ENDPOINTS (CREDIT CARD ISSUERS) ====================

@app.post("/credit-card-issuers", response_model=CreditCardIssuerRead, status_code=status.HTTP_201_CREATED)
def create_credit_card_issuer(
    payload: CreditCardIssuerCreate,
    db: Session = Depends(get_db),
    admin: User = Depends(require_admin),
):
    """Создать запись в справочнике эмитентов кредитных карт."""
    existing = db.scalar(select(CreditCardIssuer).where(CreditCardIssuer.name == payload.name))
    if existing:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Эмитент с таким именем уже существует.")
    issuer = CreditCardIssuer(**payload.model_dump())
    db.add(issuer)
    db.flush()
    _log_action(
        db,
        "credit_card_issuers.create",
        actor_user_id=admin.id,
        target_user_id=None,
        details=f"issuer_id={issuer.id}, name={issuer.name}",
    )
    db.commit()
    db.refresh(issuer)
    return issuer


@app.get("/credit-card-issuers", response_model=list[CreditCardIssuerRead])
def list_credit_card_issuers(
    db: Session = Depends(get_db),
    is_active: bool | None = Query(default=None),
):
    """Получить список эмитентов кредитных карт."""
    stmt = select(CreditCardIssuer)
    if is_active is not None:
        stmt = stmt.where(CreditCardIssuer.is_active == is_active)
    return db.scalars(stmt.order_by(CreditCardIssuer.name.asc())).all()


@app.patch("/credit-card-issuers/{issuer_id}", response_model=CreditCardIssuerRead)
def update_credit_card_issuer(
    issuer_id: int,
    payload: CreditCardIssuerUpdate,
    db: Session = Depends(get_db),
    admin: User = Depends(require_admin),
):
    """Обновить запись в справочнике эмитентов кредитных карт."""
    issuer = db.scalar(select(CreditCardIssuer).where(CreditCardIssuer.id == issuer_id))
    if not issuer:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Эмитент не найден.")
    
    updates = payload.model_dump(exclude_unset=True)
    if not updates:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Нет данных для обновления.")
    
    if "name" in updates:
        existing = db.scalar(select(CreditCardIssuer).where(CreditCardIssuer.name == updates["name"], CreditCardIssuer.id != issuer.id))
        if existing:
            raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Эмитент с таким именем уже существует.")
        issuer.name = updates["name"]
    if "description" in updates:
        issuer.description = updates["description"]
    if "is_active" in updates:
        issuer.is_active = updates["is_active"]
    
    _log_action(
        db,
        "credit_card_issuers.update",
        actor_user_id=admin.id,
        target_user_id=None,
        details=f"issuer_id={issuer.id}, fields={','.join(updates.keys())}",
    )
    db.commit()
    db.refresh(issuer)
    return issuer


@app.delete("/credit-card-issuers/{issuer_id}")
def delete_credit_card_issuer(issuer_id: int, db: Session = Depends(get_db), admin: User = Depends(require_admin)):
    """Удалить запись из справочника эмитентов кредитных карт."""
    issuer = db.scalar(select(CreditCardIssuer).where(CreditCardIssuer.id == issuer_id))
    if not issuer:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Эмитент не найден.")
    
    _log_action(
        db,
        "credit_card_issuers.delete",
        actor_user_id=admin.id,
        target_user_id=None,
        details=f"issuer_id={issuer_id}, name={issuer.name}",
    )
    db.delete(issuer)
    db.commit()
    return {"message": "Эмитент удалён."}


# ==================== DEBT HISTORY ENDPOINTS ====================

@app.get("/debt-history", response_model=list[CreditorDebtHistory])
def get_debt_history(
    creditor: str | None = Query(default=None, description="Фильтр по конкретному кредитору"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Получить историю изменения долгов по всем кредиторам или по конкретному кредитору.
    Возвращает полную историю с датами и суммами задолженностей.
    """
    stmt = select(DebtHistory).order_by(DebtHistory.creditor, DebtHistory.record_date)
    
    if creditor:
        stmt = stmt.where(DebtHistory.creditor == creditor)
    
    records = db.scalars(stmt).all()
    
    # Группируем данные по кредиторам
    from collections import defaultdict
    creditors_data = defaultdict(list)
    
    for record in records:
        creditors_data[record.creditor].append(record)
    
    result = []
    for creditor_name, history_records in creditors_data.items():
        amounts = [r.amount for r in history_records]
        result.append(CreditorDebtHistory(
            creditor=creditor_name,
            history=[DebtHistoryRead.model_validate(r) for r in history_records],
            current_amount=amounts[-1] if amounts else 0.0,
            min_amount=min(amounts) if amounts else 0.0,
            max_amount=max(amounts) if amounts else 0.0,
        ))
    
    return result


@app.get("/debt-history/{creditor_name}", response_model=CreditorDebtHistory)
def get_debt_history_by_creditor(
    creditor_name: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Получить историю изменения долга по конкретному кредитору.
    """
    stmt = select(DebtHistory).where(
        DebtHistory.creditor == creditor_name
    ).order_by(DebtHistory.record_date)
    
    records = db.scalars(stmt).all()
    
    if not records:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, 
            detail=f"История по кредитору '{creditor_name}' не найдена"
        )
    
    amounts = [r.amount for r in records]
    return CreditorDebtHistory(
        creditor=creditor_name,
        history=[DebtHistoryRead.model_validate(r) for r in records],
        current_amount=amounts[-1] if amounts else 0.0,
        min_amount=min(amounts) if amounts else 0.0,
        max_amount=max(amounts) if amounts else 0.0,
    )


# ==================== XLSX INTERPRETER API ====================

@app.post("/xlsx/upload", response_model=XLSXFileInfo)
async def upload_xlsx(
    file: UploadFile = File(..., description="XLSX файл для загрузки"),
    current_user: User = Depends(get_current_user),
):
    """
    Загрузка XLSX файла и получение информации о нём
    """
    if not file.filename.lower().endswith('.xlsx'):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Файл должен быть в формате XLSX"
        )
    
    try:
        contents = await file.read()
        xlsx_file = pd.ExcelFile(contents)
        sheet_names = xlsx_file.sheet_names
        
        return XLSXFileInfo(
            filename=file.filename,
            sheet_count=len(sheet_names),
            sheets=sheet_names
        )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Ошибка при чтении файла: {str(e)}"
        )


@app.post("/xlsx/read", response_model=XLSXSheetData)
async def read_xlsx_sheet(
    file: UploadFile = File(..., description="XLSX файл"),
    sheet_name: str | None = Query(default=None, description="Имя листа (по умолчанию первый)"),
    head: int | None = Query(default=None, description="Количество строк для чтения"),
    current_user: User = Depends(get_current_user),
):
    """
    Чтение данных из указанного листа XLSX файла
    """
    if not file.filename.lower().endswith('.xlsx'):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Файл должен быть в формате XLSX"
        )
    
    try:
        contents = await file.read()
        df = pd.read_excel(contents, sheet_name=sheet_name, nrows=head)
        
        # Преобразуем данные в JSON-сериализуемый формат
        data = []
        for _, row in df.iterrows():
            row_dict = {}
            for col, value in row.items():
                if pd.isna(value):
                    row_dict[str(col)] = None
                elif isinstance(value, (pd.Timestamp, datetime)):
                    row_dict[str(col)] = value.isoformat()
                elif isinstance(value, (int, float, str)):
                    row_dict[str(col)] = value
                else:
                    row_dict[str(col)] = str(value)
            data.append(row_dict)
        
        return XLSXSheetData(
            name=sheet_name or df.sheet_name if hasattr(df, 'sheet_name') else "Sheet1",
            columns=[str(col) for col in df.columns],
            data=data
        )
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Лист не найден: {str(e)}"
        )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Ошибка при чтении файла: {str(e)}"
        )


@app.post("/xlsx/to-json")
async def xlsx_to_json(
    file: UploadFile = File(..., description="XLSX файл"),
    sheet_name: str | None = Query(default=None, description="Имя листа"),
    current_user: User = Depends(get_current_user),
):
    """
    Конвертация XLSX листа в JSON формат
    """
    if not file.filename.lower().endswith('.xlsx'):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Файл должен быть в формате XLSX"
        )
    
    try:
        contents = await file.read()
        df = pd.read_excel(contents, sheet_name=sheet_name)
        
        # Преобразуем NaN в None для JSON сериализации
        data = df.where(pd.notna(df), None).to_dict(orient='records')
        
        # Обработка дат и других специальных типов
        for row in data:
            for key, value in row.items():
                if isinstance(value, (pd.Timestamp, datetime)):
                    row[key] = value.isoformat()
        
        return Response(
            content=json.dumps(data, indent=2, default=str),
            media_type="application/json"
        )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Ошибка при конвертации: {str(e)}"
        )


@app.post("/xlsx/to-csv")
async def xlsx_to_csv(
    file: UploadFile = File(..., description="XLSX файл"),
    sheet_name: str | None = Query(default=None, description="Имя листа"),
    current_user: User = Depends(get_current_user),
):
    """
    Конвертация XLSX листа в CSV формат
    """
    if not file.filename.lower().endswith('.xlsx'):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Файл должен быть в формате XLSX"
        )
    
    try:
        contents = await file.read()
        df = pd.read_excel(contents, sheet_name=sheet_name)
        
        csv_buffer = StringIO()
        df.to_csv(csv_buffer, index=False, encoding='utf-8-sig')
        csv_content = csv_buffer.getvalue()
        
        filename = Path(file.filename).stem + ".csv"
        
        return Response(
            content=csv_content,
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Ошибка при конвертации: {str(e)}"
        )


@app.get("/xlsx/info/{filename:path}", response_model=XLSXFileInfo)
async def get_xlsx_info(
    filename: str,
    current_user: User = Depends(get_current_user),
):
    """
    Получение информации о XLSX файле на сервере
    """
    file_path = Path(filename)
    if not file_path.is_absolute():
        file_path = Path.cwd() / file_path
    
    if not file_path.exists():
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Файл не найден"
        )
    
    if not file_path.suffix.lower() == '.xlsx':
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Файл должен быть в формате XLSX"
        )
    
    try:
        xlsx_file = pd.ExcelFile(file_path)
        sheet_names = xlsx_file.sheet_names
        
        return XLSXFileInfo(
            filename=file_path.name,
            sheet_count=len(sheet_names),
            sheets=sheet_names
        )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Ошибка при чтении файла: {str(e)}"
        )


@app.get("/xlsx/read/{filename:path}", response_model=XLSXSheetData)
async def read_server_xlsx(
    filename: str,
    sheet_name: str | None = Query(default=None, description="Имя листа"),
    head: int | None = Query(default=None, description="Количество строк"),
    current_user: User = Depends(get_current_user),
):
    """
    Чтение данных из XLSX файла на сервере
    """
    file_path = Path(filename)
    if not file_path.is_absolute():
        file_path = Path.cwd() / file_path
    
    if not file_path.exists():
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Файл не найден"
        )
    
    if not file_path.suffix.lower() == '.xlsx':
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Файл должен быть в формате XLSX"
        )
    
    try:
        df = pd.read_excel(file_path, sheet_name=sheet_name, nrows=head)
        
        # Преобразуем данные в JSON-сериализуемый формат
        data = []
        for _, row in df.iterrows():
            row_dict = {}
            for col, value in row.items():
                if pd.isna(value):
                    row_dict[str(col)] = None
                elif isinstance(value, (pd.Timestamp, datetime)):
                    row_dict[str(col)] = value.isoformat()
                elif isinstance(value, (int, float, str)):
                    row_dict[str(col)] = value
                else:
                    row_dict[str(col)] = str(value)
            data.append(row_dict)
        
        return XLSXSheetData(
            name=sheet_name or "Sheet1",
            columns=[str(col) for col in df.columns],
            data=data
        )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Ошибка при чтении файла: {str(e)}"
        )


class XLSXImportResult(BaseModel):
    """Результат импорта данных из XLSX"""
    filename: str
    sheet_name: str
    total_rows: int
    inserted: int
    updated: int
    skipped: int
    errors: List[Dict[str, Any]] = []


def _map_debt_from_row(row: Dict[str, Any], user_id: int) -> Dict[str, Any]:
    """Маппинг строки данных в модель Debt"""
    # Поддержка различных названий колонок
    creditor_map = {'creditor_name', 'creditor', 'кредитор', 'название_кредитора'}
    principal_map = {'principal_amount', 'principal', 'сумма_долга', 'основной_долг', 'тело_долга'}
    start_date_map = {'start_date', 'date', 'дата_начала', 'дата'}
    planned_date_map = {'planned_payoff_date', 'planned_date', 'дата_погашения', 'плановая_дата'}
    interest_map = {'interest_rate', 'interest', 'ставка', 'процентная_ставка'}
    comment_map = {'comment', 'comments', 'комментарий', 'примечание'}
    balance_map = {'current_balance', 'balance', 'текущий_баланс', 'остаток'}
    
    result = {'user_id': user_id}
    
    for key, value in row.items():
        key_lower = key.lower().strip() if key else ''
        
        if key_lower in creditor_map:
            result['creditor_name'] = str(value) if value else None
        elif key_lower in principal_map:
            result['principal_amount'] = _coerce_float(value)
        elif key_lower in start_date_map:
            result['start_date'] = _coerce_date(value)
        elif key_lower in planned_date_map:
            result['planned_payoff_date'] = _coerce_date(value)
        elif key_lower in interest_map:
            result['interest_rate'] = _coerce_float(value)
        elif key_lower in comment_map:
            result['comment'] = str(value) if value else None
        elif key_lower in balance_map:
            result['current_balance'] = _coerce_float(value)
    
    # Установка значений по умолчанию
    if 'current_balance' not in result and 'principal_amount' in result:
        result['current_balance'] = result['principal_amount']
    
    return result


def _map_income_from_row(row: Dict[str, Any], user_id: int) -> Dict[str, Any]:
    """Маппинг строки данных в модель Income"""
    amount_map = {'amount', 'сумма', 'доход'}
    date_map = {'income_date', 'date', 'дата_дохода', 'дата'}
    category_map = {'category', 'категория'}
    description_map = {'description', 'desc', 'описание', 'комментарий'}
    is_actual_map = {'is_actual', 'actual', 'актуальный'}
    
    result = {'user_id': user_id}
    
    for key, value in row.items():
        key_lower = key.lower().strip() if key else ''
        
        if key_lower in amount_map:
            result['amount'] = _coerce_float(value)
        elif key_lower in date_map:
            result['income_date'] = _coerce_date(value)
        elif key_lower in category_map:
            cat_value = str(value).lower() if value else None
            if cat_value:
                for cat in IncomeCategory:
                    if cat.value == cat_value or cat.name.lower() == cat_value:
                        result['category'] = cat
                        break
        elif key_lower in description_map:
            result['description'] = str(value) if value else None
        elif key_lower in is_actual_map:
            result['is_actual'] = bool(value) if value is not None else False
    
    return result


def _map_expense_from_row(row: Dict[str, Any], user_id: int) -> Dict[str, Any]:
    """Маппинг строки данных в модель Expense"""
    amount_map = {'amount', 'сумма', 'расход'}
    date_map = {'due_date', 'date', 'дата_расхода', 'дата', 'срок_оплаты'}
    category_map = {'category', 'категория'}
    description_map = {'description', 'desc', 'описание', 'комментарий'}
    is_mandatory_map = {'is_mandatory', 'mandatory', 'обязательный'}
    is_completed_map = {'is_completed', 'completed', 'выполнен'}
    
    result = {'user_id': user_id}
    
    for key, value in row.items():
        key_lower = key.lower().strip() if key else ''
        
        if key_lower in amount_map:
            result['amount'] = _coerce_float(value)
        elif key_lower in date_map:
            result['due_date'] = _coerce_date(value)
        elif key_lower in category_map:
            cat_value = str(value).lower() if value else None
            if cat_value:
                for cat in ExpenseCategory:
                    if cat.value == cat_value or cat.name.lower() == cat_value:
                        result['category'] = cat
                        break
        elif key_lower in description_map:
            result['description'] = str(value) if value else None
        elif key_lower in is_mandatory_map:
            result['is_mandatory'] = bool(value) if value is not None else False
        elif key_lower in is_completed_map:
            result['is_completed'] = bool(value) if value is not None else False
    
    return result


def _map_credit_card_from_row(row: Dict[str, Any], user_id: int) -> Dict[str, Any]:
    """Маппинг строки данных в модель CreditCard"""
    card_name_map = {'card_name', 'card', 'название_карты', 'карта'}
    grace_start_map = {'grace_start_date', 'grace_start', 'дата_начала_льготного', 'дата_начала'}
    grace_period_map = {'grace_period_days', 'grace_period', 'льготный_период', 'дней_льготного'}
    current_debt_map = {'current_debt', 'debt', 'текущий_долг', 'долг'}
    comment_map = {'comment', 'comments', 'комментарий', 'примечание'}
    
    result = {'user_id': user_id}
    
    for key, value in row.items():
        key_lower = key.lower().strip() if key else ''
        
        if key_lower in card_name_map:
            result['card_name'] = str(value) if value else None
        elif key_lower in grace_start_map:
            result['grace_start_date'] = _coerce_date(value)
        elif key_lower in grace_period_map:
            result['grace_period_days'] = int(value) if value else 30
        elif key_lower in current_debt_map:
            result['current_debt'] = _coerce_float(value) or 0.0
        elif key_lower in comment_map:
            result['comment'] = str(value) if value else None
    
    return result


@app.post("/xlsx/import", response_model=XLSXImportResult)
async def import_xlsx_to_database(
    file: UploadFile = File(..., description="XLSX файл для импорта"),
    sheet_name: str | None = Query(default=None, description="Имя листа (по умолчанию первый)"),
    entity_type: str = Query(default="auto", description="Тип данных: debt, income, expense, credit_card, auto"),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    Импорт данных из XLSX файла в базу данных
    
    Поддерживаемые типы сущностей:
    - debt: Долги (Debts)
    - income: Доходы (Incomes)
    - expense: Расходы (Expenses)
    - credit_card: Кредитные карты (CreditCards)
    - auto: Автоматическое определение по названию листа
    """
    if not file.filename.lower().endswith('.xlsx'):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Файл должен быть в формате XLSX"
        )
    
    try:
        contents = await file.read()
        
        # Используем BytesIO для правильного чтения содержимого
        from io import BytesIO
        excel_file = BytesIO(contents)
        
        # Если sheet_name не указан, читаем первый лист явно
        if sheet_name is None:
            temp_df = pd.ExcelFile(excel_file)
            sheet_name = temp_df.sheet_names[0] if temp_df.sheet_names else None
            temp_df.close()
            excel_file = BytesIO(contents)  # Пересоздаем BytesIO
        
        df = pd.read_excel(excel_file, sheet_name=sheet_name)
        
        if not hasattr(df, 'empty') or (hasattr(df, 'empty') and df.empty):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Файл не содержит данных"
            )
        
        # Автоопределение типа сущности
        actual_sheet_name = sheet_name or (df.sheet_name if hasattr(df, 'sheet_name') else "Sheet1")
        if entity_type == "auto":
            sheet_lower = actual_sheet_name.lower()
            if 'debt' in sheet_lower or 'долг' in sheet_lower or 'кредитор' in sheet_lower:
                entity_type = "debt"
            elif 'income' in sheet_lower or 'доход' in sheet_lower:
                entity_type = "income"
            elif 'expense' in sheet_lower or 'расход' in sheet_lower:
                entity_type = "expense"
            elif 'credit' in sheet_lower or 'карт' in sheet_lower:
                entity_type = "credit_card"
            else:
                # По умолчанию пытаемся определить по колонкам
                cols_lower = [str(c).lower() for c in df.columns]
                if any(c in cols_lower for c in ['creditor', 'кредитор', 'principal']):
                    entity_type = "debt"
                elif any(c in cols_lower for c in ['income', 'доход']) and 'amount' in cols_lower:
                    entity_type = "income"
                elif any(c in cols_lower for c in ['expense', 'расход']) and 'amount' in cols_lower:
                    entity_type = "expense"
                elif any(c in cols_lower for c in ['card', 'карта', 'grace']):
                    entity_type = "credit_card"
                else:
                    entity_type = "debt"  # Default
        
        result = XLSXImportResult(
            filename=file.filename,
            sheet_name=actual_sheet_name,
            total_rows=len(df),
            inserted=0,
            updated=0,
            skipped=0,
            errors=[]
        )
        
        # Преобразуем строки в словари
        rows_data = []
        for _, row in df.iterrows():
            row_dict = {}
            for col, value in row.items():
                if pd.isna(value):
                    row_dict[str(col)] = None
                elif isinstance(value, (pd.Timestamp, datetime)):
                    row_dict[str(col)] = value.isoformat()
                elif isinstance(value, (int, float, str)):
                    row_dict[str(col)] = value
                else:
                    row_dict[str(col)] = str(value)
            rows_data.append(row_dict)
        
        moderation_status = RecordStatus.APPROVED if current_user.role == UserRole.ADMIN else RecordStatus.PENDING
        
        if entity_type == "debt":
            for idx, row in enumerate(rows_data):
                try:
                    mapped = _map_debt_from_row(row, current_user.id)
                    
                    # Проверка обязательных полей
                    if not mapped.get('creditor_name') or not mapped.get('principal_amount') or not mapped.get('start_date'):
                        result.errors.append({
                            "row": idx + 1,
                            "error": "Отсутствуют обязательные поля (кредитор, сумма, дата начала)"
                        })
                        result.skipped += 1
                        continue
                    
                    # Проверка на дубликат
                    existing = db.scalar(
                        select(Debt).where(
                            Debt.user_id == current_user.id,
                            Debt.creditor_name == mapped['creditor_name'],
                            Debt.start_date == mapped['start_date']
                        )
                    )
                    
                    if existing:
                        # Обновление существующей записи
                        for key, value in mapped.items():
                            if hasattr(existing, key) and value is not None:
                                setattr(existing, key, value)
                        existing.moderation_status = moderation_status
                        if moderation_status == RecordStatus.APPROVED:
                            existing.approved_by_id = current_user.id
                            existing.approved_at = datetime.utcnow()
                        result.updated += 1
                    else:
                        # Создание новой записи
                        debt = Debt(
                            **mapped,
                            moderation_status=moderation_status,
                            approved_by_id=current_user.id if moderation_status == RecordStatus.APPROVED else None,
                            approved_at=datetime.utcnow() if moderation_status == RecordStatus.APPROVED else None
                        )
                        db.add(debt)
                        result.inserted += 1
                        
                except Exception as e:
                    result.errors.append({"row": idx + 1, "error": str(e)})
                    result.skipped += 1
            
            db.commit()
            
        elif entity_type == "income":
            for idx, row in enumerate(rows_data):
                try:
                    mapped = _map_income_from_row(row, current_user.id)
                    
                    if not mapped.get('amount') or not mapped.get('income_date'):
                        result.errors.append({
                            "row": idx + 1,
                            "error": "Отсутствуют обязательные поля (сумма, дата)"
                        })
                        result.skipped += 1
                        continue
                    
                    income = Income(
                        **mapped,
                        moderation_status=moderation_status,
                        approved_by_id=current_user.id if moderation_status == RecordStatus.APPROVED else None,
                        approved_at=datetime.utcnow() if moderation_status == RecordStatus.APPROVED else None
                    )
                    db.add(income)
                    result.inserted += 1
                    
                except Exception as e:
                    result.errors.append({"row": idx + 1, "error": str(e)})
                    result.skipped += 1
            
            db.commit()
            
        elif entity_type == "expense":
            for idx, row in enumerate(rows_data):
                try:
                    mapped = _map_expense_from_row(row, current_user.id)
                    
                    if not mapped.get('amount') or not mapped.get('due_date'):
                        result.errors.append({
                            "row": idx + 1,
                            "error": "Отсутствуют обязательные поля (сумма, дата)"
                        })
                        result.skipped += 1
                        continue
                    
                    expense = Expense(
                        **mapped,
                        moderation_status=moderation_status,
                        approved_by_id=current_user.id if moderation_status == RecordStatus.APPROVED else None,
                        approved_at=datetime.utcnow() if moderation_status == RecordStatus.APPROVED else None
                    )
                    db.add(expense)
                    result.inserted += 1
                    
                except Exception as e:
                    result.errors.append({"row": idx + 1, "error": str(e)})
                    result.skipped += 1
            
            db.commit()
            
        elif entity_type == "credit_card":
            for idx, row in enumerate(rows_data):
                try:
                    mapped = _map_credit_card_from_row(row, current_user.id)
                    
                    if not mapped.get('card_name') or not mapped.get('grace_start_date'):
                        result.errors.append({
                            "row": idx + 1,
                            "error": "Отсутствуют обязательные поля (название карты, дата начала)"
                        })
                        result.skipped += 1
                        continue
                    
                    credit_card = CreditCard(
                        **mapped,
                        moderation_status=moderation_status,
                        approved_by_id=current_user.id if moderation_status == RecordStatus.APPROVED else None,
                        approved_at=datetime.utcnow() if moderation_status == RecordStatus.APPROVED else None
                    )
                    db.add(credit_card)
                    result.inserted += 1
                    
                except Exception as e:
                    result.errors.append({"row": idx + 1, "error": str(e)})
                    result.skipped += 1
            
            db.commit()
        
        else:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Неподдерживаемый тип сущности: {entity_type}"
            )
        
        # Логирование
        _log_action(
            db,
            "xlsx.import",
            actor_user_id=current_user.id,
            details=f"file={file.filename}, sheet={actual_sheet_name}, type={entity_type}, inserted={result.inserted}"
        )
        
        return result
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Ошибка при импорте: {str(e)}"
        )
