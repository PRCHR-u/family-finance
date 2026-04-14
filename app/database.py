import os
from sqlalchemy import create_engine
from sqlalchemy.orm import declarative_base, sessionmaker

# Чтение переменных окружения
DATABASE_URL = os.getenv("DATABASE_URL", "sqlite:///./family_finance.db")

# Для SQLite важно правильно обрабатывать путь
if DATABASE_URL.startswith("sqlite"):
    engine = create_engine(DATABASE_URL, connect_args={"check_same_thread": False})
else:
    engine = create_engine(DATABASE_URL)

SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
Base = declarative_base()


def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


def init_db():
    """Инициализация базы данных и создание начальных данных"""
    from .models import User, UserRole, Creditor, CreditCardIssuer, ExpenseCategory, IncomeCategory, DebtHistory
    from .auth import get_password_hash
    from sqlalchemy.orm import Session
    
    Base.metadata.create_all(bind=engine)
    
    db = SessionLocal()
    try:
        # Создание администратора по умолчанию
        admin_username = os.getenv("ADMIN_USERNAME", "admin")
        admin_email = os.getenv("ADMIN_EMAIL", "admin@example.com")
        admin_password = os.getenv("ADMIN_PASSWORD", "ChangeMe123!")
        
        admin = db.query(User).filter(User.username == admin_username).first()
        if not admin:
            admin = User(
                username=admin_username,
                hashed_password=get_password_hash(admin_password),
                role=UserRole.ADMIN,
                is_active=True
            )
            db.add(admin)
            db.commit()
            print(f"Администратор создан: {admin_username} ({admin_email})")
        
        # Импорт данных из Excel файла ДОЛГИ.xlsx
        _import_debt_history_from_excel(db)
        
        print("База данных инициализирована успешно")
    except Exception as e:
        db.rollback()
        print(f"Ошибка инициализации БД: {e}")
        raise
    finally:
        db.close()


def _import_debt_history_from_excel(db):
    """Импортирует историю долгов из файла ДОЛГИ.xlsx"""
    import os
    from pathlib import Path
    from datetime import datetime
    from .models import DebtHistory
    
    # Проверяем, есть ли уже данные в таблице
    existing_count = db.query(DebtHistory).count()
    if existing_count > 0:
        print(f"Таблица debt_history уже содержит {existing_count} записей. Пропускаем импорт.")
        return
    
    # Путь к файлу
    project_root = Path(__file__).resolve().parent.parent
    excel_path = project_root / "ДОЛГИ.xlsx"
    
    if not excel_path.exists():
        print(f"Файл {excel_path} не найден. Пропускаем импорт истории долгов.")
        return
    
    try:
        import openpyxl
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        
        creditors = ['СБЕР', 'АЛЬФА', 'МТС1', 'МТС2', 'Т-БАНК', 'ОЛЯ', 'КРЕДИТ']
        creditor_indices = {name: idx for idx, name in enumerate(creditors, start=1)}
        
        records_to_add = []
        
        for row_idx in range(2, ws.max_row + 1):
            row_data = [cell.value for cell in ws[row_idx]]
            
            # Проверяем, есть ли дата в первой колонке
            first_cell = row_data[0]
            if isinstance(first_cell, datetime):
                record_date = first_cell.date()
                
                # Собираем данные по каждому кредитору
                for creditor_name, col_idx in creditor_indices.items():
                    if col_idx < len(row_data):
                        value = row_data[col_idx]
                        if value is not None and isinstance(value, (int, float)):
                            records_to_add.append(DebtHistory(
                                creditor=creditor_name,
                                amount=float(value),
                                record_date=record_date
                            ))
        
        if records_to_add:
            db.bulk_save_objects(records_to_add)
            db.commit()
            print(f"Импортировано {len(records_to_add)} записей истории долгов из {excel_path}")
        else:
            print("Не найдено данных для импорта в файле Excel")
            
    except ImportError:
        print("Библиотека openpyxl не установлена. Пропускаем импорт истории долгов.")
    except Exception as e:
        print(f"Ошибка при импорте истории долгов: {e}")
        raise
